"""
Monitor de Inventario Inteligente - Colsabor
Aplicación Streamlit para monitorear inventario conectado a Siigo API
"""

import io
import logging
import sys
import time
import streamlit as st
import pandas as pd
import requests
import plotly.graph_objects as go
import base64
import tomllib
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    _OPENPYXL_OK = True
except ImportError:  # pragma: no cover
    _OPENPYXL_OK = False
from supabase import create_client

try:
    from inventory_monitor.loading_messages import get_greeting, get_random_message
    from inventory_monitor import dane_survey
except ModuleNotFoundError:  # pragma: no cover
    from loading_messages import get_greeting, get_random_message  # type: ignore[no-redef]  # pragma: no cover
    import dane_survey  # type: ignore[no-redef]  # pragma: no cover

# ── Debug logger (imprime en consola/terminal donde corre Streamlit) ──────────
logging.basicConfig(
    level=logging.DEBUG,
    format="[CS-DEBUG %(asctime)s] %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stdout,
    force=True,
)
_log = logging.getLogger("colsabor")

# ============================================================================
# CONFIGURACIÓN DE CREDENCIALES SIIGO API
# ============================================================================
# URL base de la API de Siigo (fija para todos los usuarios)
SIIGO_API_BASE_URL = "https://api.siigo.com/v1"
# Access Key compartido de la empresa (mismo para todos)
SIIGO_ACCESS_KEY = "MmQzMDk0NjYtZjc3Ny00YzU0LWFmNDMtMjhiYzcxNGM5NTBhOnoyeTk5KE4uYkc="

# ============================================================================
# CONFIGURACIÓN DE SUPABASE
# ============================================================================


def _load_supabase_from_local_secrets() -> tuple[str, str]:
    """Fallback de credenciales Supabase desde secrets.toml local."""
    candidate_paths = [
        Path.cwd() / ".streamlit" / "secrets.toml",
        Path(__file__).resolve().parent / ".streamlit" / "secrets.toml",
        Path(__file__).resolve().parents[1] / ".streamlit" / "secrets.toml",
    ]
    _log.debug("[SECRETS-FALLBACK] cwd=%s", Path.cwd())
    for secrets_path in candidate_paths:
        _log.debug(
            "[SECRETS-FALLBACK] probando: %s  existe=%s",
            secrets_path,
            secrets_path.exists(),
        )
        if not secrets_path.exists():  # pragma: no cover
            continue
        try:
            data = tomllib.loads(secrets_path.read_text(encoding="utf-8"))
        except Exception as exc:  # pragma: no cover
            _log.debug("[SECRETS-FALLBACK] error leyendo %s: %s", secrets_path, exc)
            continue

        url = str(data.get("SUPABASE_URL", "") or "").strip()
        key = str(data.get("SUPABASE_KEY", "") or "").strip()
        _log.debug(
            "[SECRETS-FALLBACK] leido -> URL=%s KEY_prefix=%s",
            url[:30] if url else "VACIO",
            key[:20] if key else "VACIO",
        )
        if url and key:
            return url, key

    _log.debug(
        "[SECRETS-FALLBACK] no se encontraron credenciales en ninguna ruta"
    )  # pragma: no cover
    return "", ""  # pragma: no cover


# Anon key pública del proyecto Supabase (seguro incluir en código fuente)
_SUPABASE_URL_DEFAULT = "https://uinqrkxlkjowixmtzold.supabase.co"
_SUPABASE_KEY_DEFAULT = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InVpbnFya3hsa2pvd2l4bXR6b2xkIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzYyMTg2OTgsImV4cCI6MjA5MTc5NDY5OH0"
    ".ifGRvqwAtI-6D72_BC7uih-88boy2wcseBUEi-o_0ek"
)

_log.debug("[INIT] cargando credenciales de st.secrets...")
SUPABASE_URL = st.secrets.get("SUPABASE_URL", "")
SUPABASE_KEY = st.secrets.get("SUPABASE_KEY", "")
_log.debug(
    "[INIT] st.secrets -> URL=%s KEY_prefix=%s",
    (SUPABASE_URL or "VACIO")[:35],
    (SUPABASE_KEY or "VACIO")[:20],
)
if not SUPABASE_URL or not SUPABASE_KEY:
    _log.debug("[INIT] st.secrets vacio, usando fallback de archivo local...")
    _fallback_url, _fallback_key = _load_supabase_from_local_secrets()
    SUPABASE_URL = SUPABASE_URL or _fallback_url
    SUPABASE_KEY = SUPABASE_KEY or _fallback_key
    _log.debug(
        "[INIT] tras fallback local -> URL=%s KEY_prefix=%s",
        (SUPABASE_URL or "VACIO")[:35],
        (SUPABASE_KEY or "VACIO")[:20],
    )
if not SUPABASE_URL or not SUPABASE_KEY:  # pragma: no cover
    _log.debug(
        "[INIT] fallback a constantes hardcodeadas (anon key publica)"
    )  # pragma: no cover
    SUPABASE_URL = SUPABASE_URL or _SUPABASE_URL_DEFAULT  # pragma: no cover
    SUPABASE_KEY = SUPABASE_KEY or _SUPABASE_KEY_DEFAULT  # pragma: no cover
_log.debug(
    "[INIT] credenciales finales -> URL=%s KEY_prefix=%s KEY_len=%d",
    SUPABASE_URL[:40],
    SUPABASE_KEY[:20],
    len(SUPABASE_KEY),
)

# Nombres de las tablas en Supabase
TABLE_INVENTARIO = "user_inventory"
TABLE_SIIGO_CACHE = "siigo_products_cache"
COL_TZ = ZoneInfo("America/Bogota")


def now_colombia() -> datetime:
    """Retorna fecha/hora actual en zona horaria de Colombia."""
    return datetime.now(COL_TZ)

# ============================================================================
# USUARIOS AUTORIZADOS (whitelist)
# ============================================================================
ALLOWED_EMAILS = {
    "dirtec@colsabor.com.co",
    "gerencia@colsabor.com.co",
    "samuelrestrepodev@gmail.com",
}

# ============================================================================
# CONFIGURACIÓN DE LA PÁGINA
# ============================================================================
st.set_page_config(
    page_title="Monitor de Inventario · Colsabor",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ============================================================================
# FUNCIONES DE AUTENTICACIÓN Y API SIIGO
# ============================================================================


def autenticar_siigo(username: str, access_key: str) -> dict:
    """
    Autentica con la API de Siigo y obtiene el token de acceso.

    Args:
        username: Usuario de Siigo
        access_key: Clave de acceso de Siigo

    Returns:
        dict: Respuesta con el token o error
    """
    url = "https://api.siigo.com/auth"

    headers = {"Content-Type": "application/json", "Partner-Id": "ColsaborApp"}

    payload = {"username": username, "access_key": access_key}

    try:
        response = requests.post(url, json=payload, headers=headers, timeout=30)

        if response.status_code == 200:
            return {"success": True, "token": response.json().get("access_token")}
        else:
            return {
                "success": False,
                "error": f"Error de autenticación: {response.status_code} - {response.text}",
            }
    except requests.exceptions.RequestException as e:
        return {"success": False, "error": f"Error de conexión: {str(e)}"}


def obtener_todos_los_productos_siigo(
    token: str, referencias_requeridas: list[str] | None = None
) -> dict:
    """
    Obtiene productos de Siigo. Si referencias_requeridas se proporciona, solo obtiene esos.
    Si no, obtiene TODOS con paginación.

    Args:
        token: Token de autenticación
        referencias_requeridas: Lista de códigos de producto a obtener (para optimización)

    Returns:
        dict: Lista de productos
    """
    url = f"{SIIGO_API_BASE_URL}/products"

    headers = {
        "Authorization": token,
        "Content-Type": "application/json",
        "Partner-Id": "ColsaborApp",
    }

    todos_productos = []

    try:
        # ── Opción optimizada: buscar solo referencias requeridas ────
        if (
            referencias_requeridas and len(referencias_requeridas) > 0
        ):  # pragma: no cover
            # Hacer búsqueda individual para cada referencia (más rápido que descargar todos)
            for ref in referencias_requeridas:  # pragma: no cover
                params = {"code": str(ref).strip()}  # pragma: no cover
                try:  # pragma: no cover
                    response = requests.get(
                        url, headers=headers, params=params, timeout=30
                    )  # pragma: no cover
                    if response.status_code == 200:  # pragma: no cover
                        data = response.json()  # pragma: no cover
                        productos_pagina = []  # pragma: no cover
                        if isinstance(data, list):  # pragma: no cover
                            productos_pagina = data  # pragma: no cover
                        elif (
                            isinstance(data, dict) and "results" in data
                        ):  # pragma: no cover
                            productos_pagina = data["results"]  # pragma: no cover
                        todos_productos.extend(productos_pagina)  # pragma: no cover
                except requests.exceptions.RequestException:  # pragma: no cover
                    pass  # pragma: no cover
            return {
                "success": True,
                "data": todos_productos,
                "total": len(todos_productos),
            }  # pragma: no cover

        # ── Opción estándar: obtener TODOS los productos con paginación ────
        page = 1
        page_size = 100

        while True:
            params = {"page": page, "page_size": page_size}

            response = requests.get(url, headers=headers, params=params, timeout=60)

            if response.status_code == 200:
                data = response.json()

                # Manejar diferentes formatos de respuesta
                productos_pagina = []
                if isinstance(data, list):
                    productos_pagina = data
                elif isinstance(data, dict) and "results" in data:
                    productos_pagina = data["results"]

                # Si no hay más productos, salir del loop
                if not productos_pagina or len(productos_pagina) == 0:
                    break

                todos_productos.extend(productos_pagina)

                # Si obtuvimos menos productos que el page_size, es la última página
                if len(productos_pagina) < page_size:
                    break

                page += 1

            else:
                return {
                    "success": False,
                    "error": f"Error al obtener productos: {response.status_code} - {response.text}",
                }

        return {"success": True, "data": todos_productos, "total": len(todos_productos)}

    except requests.exceptions.RequestException as e:
        return {"success": False, "error": f"Error de conexión: {str(e)}"}


# ============================================================================
# FUNCIONES DE SUPABASE (BASE DE DATOS)
# ============================================================================


def get_supabase_client():
    """Inicializa el cliente de Supabase."""
    _log.debug(
        "[SUPABASE-CLIENT] URL=%s KEY_prefix=%s KEY_len=%d",
        (SUPABASE_URL or "VACIO")[:40],
        (SUPABASE_KEY or "VACIO")[:20],
        len(SUPABASE_KEY or ""),
    )
    if not SUPABASE_URL or not SUPABASE_KEY:
        _log.debug("[SUPABASE-CLIENT] credenciales vacias -> retornando None")
        st.warning(
            "⚠️ No se encontraron credenciales de Supabase en secrets. Los datos no se guardarán."
        )
        return None
    try:
        _log.debug("[SUPABASE-CLIENT] llamando create_client...")
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        _log.debug("[SUPABASE-CLIENT] create_client OK")
        return supabase
    except Exception as e:
        _log.debug("[SUPABASE-CLIENT] create_client ERROR: %s", e)
        st.error(f"Error al conectar con Supabase: {str(e)}")
        return None


def guardar_inventario_excel(usuario_email: str, df_excel: pd.DataFrame):
    """Guarda el inventario Excel del usuario en Supabase."""
    try:
        supabase = get_supabase_client()
        if not supabase:
            return False

        records = df_excel.to_dict(orient="records")
        data_to_insert = {
            "usuario_email": usuario_email,
            "data": records,
            "updated_at": now_colombia().isoformat(),
        }

        supabase.table(TABLE_INVENTARIO).upsert(
            data_to_insert, on_conflict="usuario_email"
        ).execute()
        return True
    except Exception as e:
        st.error(f"Error al guardar en Supabase: {str(e)}")
        return False


def cargar_inventario_guardado(usuario_email: str):
    """Carga el inventario guardado del usuario desde Supabase."""
    try:
        supabase = get_supabase_client()
        if not supabase:
            return None

        response = (
            supabase.table(TABLE_INVENTARIO)
            .select("data")
            .eq("usuario_email", usuario_email)
            .execute()
        )
        if not response.data:
            return None

        records = response.data[0]["data"]
        df = pd.DataFrame(records)

        if "inventario_minimo" in df.columns:
            df["inventario_minimo"] = pd.to_numeric(
                df["inventario_minimo"], errors="coerce"
            )

        return df
    except Exception as e:
        st.error(f"Error al cargar desde Supabase: {str(e)}")
        return None


def guardar_productos_siigo(productos_siigo: list):
    """Guarda los productos de Siigo en Supabase (caché compartido)."""
    try:
        supabase = get_supabase_client()
        if not supabase:
            return False

        data_to_insert = {
            "id": 1,
            "data": productos_siigo,
            "updated_at": now_colombia().isoformat(),
        }

        supabase.table(TABLE_SIIGO_CACHE).upsert(
            data_to_insert, on_conflict="id"
        ).execute()
        return True
    except Exception as e:
        st.warning(f"No se pudieron guardar productos de Siigo: {str(e)}")
        return False


def cargar_productos_siigo_guardados():
    """Carga los productos de Siigo guardados desde Supabase."""
    try:
        supabase = get_supabase_client()
        if not supabase:
            return None

        response = supabase.table(TABLE_SIIGO_CACHE).select("*").eq("id", 1).execute()
        if not response.data:
            return None

        row = response.data[0]

        updated_at_str = str(row["updated_at"])
        if updated_at_str.endswith("Z"):
            updated_at_str = updated_at_str.replace("Z", "+00:00")

        updated_at = datetime.fromisoformat(updated_at_str)
        if updated_at.tzinfo is None:
            updated_at = updated_at.replace(tzinfo=COL_TZ)
        updated_at = updated_at.astimezone(COL_TZ)

        horas_transcurridas = (now_colombia() - updated_at).total_seconds() / 3600
        if horas_transcurridas > 24:
            return None

        productos_siigo = row["data"]
        df_siigo = procesar_productos_siigo(productos_siigo)
        return (df_siigo, None, updated_at)
    except Exception:
        return None


# ============================================================================
# FUNCIONES DE PROCESAMIENTO DE DATOS
# ============================================================================


def cargar_inventario_minimo_supabase() -> pd.DataFrame:
    """
    Carga el inventario mínimo desde la tabla inventario_minimo en Supabase.
    """
    try:
        supabase = get_supabase_client()
        if not supabase:
            return None
        response = supabase.table("inventario_minimo").select("*").execute()
        if not response.data:
            return None
        df = pd.DataFrame(response.data)
        df = df.rename(
            columns={
                "codigo": "referencia",
                "nombre": "nombre",
                "inv_minimo": "inventario_minimo",
            }
        )
        df["referencia"] = df["referencia"].astype(str).str.strip()
        df["nombre"] = df["nombre"].astype(str).str.strip()
        df["inventario_minimo"] = pd.to_numeric(
            df["inventario_minimo"], errors="coerce"
        ).fillna(0)
        return df
    except Exception as e:
        st.error(f"Error al cargar inventario mínimo desde Supabase: {str(e)}")
        return None


def procesar_productos_siigo(productos: list) -> pd.DataFrame:
    """
    Procesa la lista de productos de Siigo a un DataFrame.

    Args:
        productos: Lista de productos de la API

    Returns:
        pd.DataFrame: DataFrame con productos procesados
    """
    datos = []

    for producto in productos:
        # Extraer información relevante
        referencia = producto.get("code", "")
        nombre = producto.get("name", "")

        # Si no tiene código o nombre, saltar
        if not referencia or not nombre:
            continue

        # El stock puede venir en diferentes estructuras según la API
        stock_actual = 0

        # Intentar obtener stock de diferentes ubicaciones posibles
        if "available_quantity" in producto:
            stock_actual = float(producto["available_quantity"])
        elif "stock" in producto:
            stock_actual = float(producto["stock"])
        elif "warehouses" in producto and isinstance(producto["warehouses"], list):
            # Sumar stock de todas las bodegas
            for bodega in producto["warehouses"]:
                stock_actual += float(bodega.get("quantity", 0))

        datos.append(
            {
                "referencia_siigo": str(referencia).strip(),
                "nombre_siigo": nombre.strip(),
                "stock_actual": stock_actual,
            }
        )

    # Crear DataFrame vacío con las columnas correctas si no hay datos
    if len(datos) == 0:
        return pd.DataFrame(
            columns=["referencia_siigo", "nombre_siigo", "stock_actual"]
        )

    return pd.DataFrame(datos)


def cruzar_inventarios(df_excel: pd.DataFrame, df_siigo: pd.DataFrame) -> pd.DataFrame:
    """
    Cruza los datos del Excel con los de Siigo y determina estado.
    Primero intenta matcheo por referencia exacta, luego por nombre normalizado.

    Args:
        df_excel: DataFrame del Excel
        df_siigo: DataFrame de Siigo

    Returns:
        pd.DataFrame: DataFrame con el cruce y estado
    """
    # Realizar merge por referencia exacta (primero intento)
    df_cruzado = df_excel.merge(
        df_siigo, left_on="referencia", right_on="referencia_siigo", how="left"
    )

    # Normalizar nombres para búsqueda secundaria
    df_siigo_aux = df_siigo.copy()
    df_siigo_aux["nombre_normalizado"] = (
        df_siigo_aux["nombre_siigo"]
        .astype(str)
        .str.upper()
        .str.strip()
        .str.replace(r'\s+', ' ', regex=True)
    )
    
    # Para productos no encontrados, intentar matcheo por nombre normalizado
    no_encontrados_mask = df_cruzado["referencia_siigo"].isna()
    if no_encontrados_mask.any():
        df_cruzado["nombre_normalizado"] = (
            df_cruzado["nombre"]
            .astype(str)
            .str.upper()
            .str.strip()
            .str.replace(r'\s+', ' ', regex=True)
        )
        
        # Buscar matches por nombre para productos faltantes
        for idx in df_cruzado[no_encontrados_mask].index:  # pragma: no cover
            nombre_norm = df_cruzado.loc[idx, "nombre_normalizado"]  # pragma: no cover
            match = df_siigo_aux[
                df_siigo_aux["nombre_normalizado"] == nombre_norm
            ]  # pragma: no cover
            if len(match) > 0:  # pragma: no cover
                # Usar el primer match encontrado
                match_idx = match.index[0]  # pragma: no cover
                df_cruzado.loc[idx, "referencia_siigo"] = df_siigo_aux.loc[  # pragma: no cover
                    match_idx, "referencia_siigo"
                ]  # pragma: no cover
                df_cruzado.loc[idx, "nombre_siigo"] = df_siigo_aux.loc[  # pragma: no cover
                    match_idx, "nombre_siigo"
                ]  # pragma: no cover
                df_cruzado.loc[idx, "stock_actual"] = df_siigo_aux.loc[  # pragma: no cover
                    match_idx, "stock_actual"
                ]  # pragma: no cover

    # Marcar productos no encontrados en Siigo
    df_cruzado["encontrado_en_siigo"] = df_cruzado["referencia_siigo"].notna()
    df_cruzado["stock_actual"] = df_cruzado["stock_actual"].fillna(0)

    # Calcular diferencia
    df_cruzado["diferencia"] = (
        df_cruzado["stock_actual"] - df_cruzado["inventario_minimo"]
    )

    # Determinar estado
    def determinar_estado(row):
        if not row["encontrado_en_siigo"]:
            return "⚠️ No encontrado en Siigo"
        elif row["stock_actual"] < row["inventario_minimo"]:
            return "🔴 Crítico"
        elif row["stock_actual"] <= row["inventario_minimo"] * 1.2:
            return "🟡 Bajo"
        else:
            return "🟢 OK"

    df_cruzado["estado"] = df_cruzado.apply(determinar_estado, axis=1)

    # Seleccionar y ordenar columnas finales
    columnas_finales = [
        "referencia",
        "nombre",
        "inventario_minimo",
        "stock_actual",
        "diferencia",
        "estado",
    ]

    df_resultado = df_cruzado[columnas_finales].copy()
    df_resultado.columns = [
        "Referencia",
        "Nombre",
        "Mínimo (g)",
        "Stock Actual",
        "Diferencia",
        "Estado",
    ]

    return df_resultado


# ============================================================================
# FUNCIONES DE EXPORTACIÓN E IMPRESIÓN
# ============================================================================


def generar_html_impresion(df: pd.DataFrame, titulo: str = "Lista de Faltantes") -> str:
    """
    Genera HTML formateado para impresión.

    Args:
        df: DataFrame con los datos a imprimir
        titulo: Título del reporte

    Returns:
        str: HTML formateado
    """
    fecha_actual = now_colombia().strftime("%d/%m/%Y %H:%M")

    # Generar filas de la tabla
    filas_html = ""
    for _, row in df.iterrows():
        estado_class = ""
        if "Crítico" in str(row["Estado"]):
            estado_class = "background-color: #ffcccc;"
        elif "Bajo" in str(row["Estado"]):
            estado_class = "background-color: #fff3cd;"

        filas_html += f"""
        <tr style="{estado_class}">
            <td>{row['Referencia']}</td>
            <td>{row['Nombre']}</td>
            <td style="text-align: right;">{row['Mínimo (g)']:,.0f}</td>
            <td style="text-align: right;">{row['Stock Actual']:,.0f}</td>
            <td style="text-align: right;">{row['Diferencia']:,.0f}</td>
            <td>{row['Estado']}</td>
        </tr>
        """

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>{titulo}</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 20px;
                font-size: 12px;
            }}
            h1 {{
                color: #1E88E5;
                text-align: center;
                margin-bottom: 5px;
            }}
            .fecha {{
                text-align: center;
                color: #666;
                margin-bottom: 20px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 20px;
            }}
            th, td {{
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }}
            th {{
                background-color: #1E88E5;
                color: white;
            }}
            tr:nth-child(even) {{
                background-color: #f9f9f9;
            }}
            .footer {{
                margin-top: 20px;
                text-align: center;
                color: #666;
                font-size: 10px;
            }}
            @media print {{
                body {{ margin: 0; }}
                .no-print {{ display: none; }}
            }}
        </style>
    </head>
    <body>
        <h1>📦 {titulo}</h1>
        <p class="fecha">Generado el: {fecha_actual}</p>

        <table>
            <thead>
                <tr>
                    <th>Referencia</th>
                    <th>Nombre</th>
                    <th>Mínimo (g)</th>
                    <th>Stock Actual</th>
                    <th>Diferencia</th>
                    <th>Estado</th>
                </tr>
            </thead>
            <tbody>
                {filas_html}
            </tbody>
        </table>

        <div class="footer">
            <p>Colsabor - Sistema de Monitor de Inventario</p>
            <p>Total de productos listados: {len(df)}</p>
        </div>

        <script>
            // Auto-abrir diálogo de impresión
            window.onload = function() {{
                window.print();
            }}
        </script>
    </body>
    </html>
    """

    return html


def generar_csv_descarga(df: pd.DataFrame) -> bytes:
    """Genera archivo CSV para descarga."""
    return df.to_csv(index=False).encode("utf-8")


def generar_excel_tabla_descarga(
    df: pd.DataFrame,
    *,
    sheet_title: str = "Datos",
    table_display_name: str = "TablaInventario",
) -> bytes:
    """
    Genera un .xlsx con tabla de Excel nativa (equivalente a Insertar > Tabla),
    encabezado con estilo y anchos de columna razonables.
    Si openpyxl no está disponible retorna CSV plano como fallback.
    """
    if not _OPENPYXL_OK:  # pragma: no cover
        return df.to_csv(index=False).encode("utf-8")
    df_out = df.copy()
    wb = Workbook()
    ws = wb.active
    safe_title = sheet_title[:31].replace("/", "-")
    ws.title = safe_title

    nrows, ncols = len(df_out) + 1, max(1, len(df_out.columns))
    header_fill = PatternFill("solid", fgColor="1F6F4F")
    header_font = Font(color="FFFFFF", bold=True)
    thin_align = Alignment(vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(df_out.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = thin_align

    for r, row in enumerate(df_out.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    last_col = get_column_letter(ncols)
    last_row = len(df_out) + 1
    tab_ref = f"A1:{last_col}{last_row}"
    tab = Table(displayName=table_display_name[:255], ref=tab_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    for i, col in enumerate(df_out.columns, 1):
        letter = get_column_letter(i)
        sample = [str(col)] + [str(x) for x in df_out.iloc[:, i - 1].head(50).tolist()]
        w = min(48, max(10, max(len(s) for s in sample) + 2))
        ws.column_dimensions[letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================================
# INTERFAZ PRINCIPAL
# ============================================================================

_VARS_LIGHT = """
  --bg-base: #f0f4fa;
  --bg-surface: rgba(255,255,255,0.96);
  --bg-surface-hover: rgba(255,255,255,1);
  --bg-glass: rgba(255,255,255,0.80);
  --border-subtle: rgba(148,163,184,0.25);
  --border-strong: rgba(100,116,139,0.40);
  --text-primary: #07111e;
  --text-secondary: #334155;
  --text-muted: #64748b;
    --placeholder-color: #334155;
  --accent: #2563eb;
  --accent-hover: #1d4ed8;
  --accent-light: #dbeafe;
  --accent-glow: rgba(37,99,235,0.20);
  --cyan: #0891b2;
  --red: #dc2626; --red-bg: #fef2f2; --red-border: rgba(220,38,38,0.22);
  --amber: #d97706; --amber-bg: #fffbeb; --amber-border: rgba(217,119,6,0.22);
  --green: #059669; --green-bg: #ecfdf5; --green-border: rgba(5,150,105,0.22);
  --slate: #64748b; --slate-bg: #f8fafc; --slate-border: rgba(100,116,139,0.22);
  --shadow-sm: 0 1px 3px rgba(0,0,0,0.07);
  --shadow-md: 0 4px 20px rgba(0,0,0,0.09);
  --shadow-lg: 0 16px 48px rgba(0,0,0,0.12);
  --shadow-glow: 0 0 0 1px rgba(37,99,235,0.07), 0 8px 32px rgba(37,99,235,0.12);
  --pattern: rgba(37,99,235,0.028);
  --nav-bg: rgba(240,244,250,0.96);
  --nav-border: rgba(148,163,184,0.30);
  --input-bg: rgba(255,255,255,0.96);
  --toggle-bg: rgba(0,0,0,0.05);
  --card-stripe-blue: linear-gradient(90deg,#2563eb,#0891b2);
  --card-stripe-red: linear-gradient(90deg,#dc2626,#ef4444);
  --card-stripe-amber: linear-gradient(90deg,#d97706,#f59e0b);
  --card-stripe-green: linear-gradient(90deg,#059669,#34d399);
"""

_VARS_DARK = """
  --bg-base: #030712;
  --bg-surface: rgba(7,16,40,0.92);
  --bg-surface-hover: rgba(10,22,52,0.96);
  --bg-glass: rgba(5,11,28,0.82);
  --border-subtle: rgba(37,99,235,0.16);
  --border-strong: rgba(59,130,246,0.28);
  --text-primary: #f0f6ff;
  --text-secondary: #cbd5e1;
  --text-muted: #94a3b8;
    --placeholder-color: #94a3b8;
  --accent: #3b82f6;
  --accent-hover: #60a5fa;
  --accent-light: rgba(59,130,246,0.14);
  --accent-glow: rgba(59,130,246,0.30);
  --cyan: #22d3ee;
  --red: #f87171; --red-bg: rgba(248,113,113,0.09); --red-border: rgba(248,113,113,0.25);
  --amber: #fbbf24; --amber-bg: rgba(251,191,36,0.09); --amber-border: rgba(251,191,36,0.25);
  --green: #34d399; --green-bg: rgba(52,211,153,0.09); --green-border: rgba(52,211,153,0.25);
  --slate: #64748b; --slate-bg: rgba(100,116,139,0.08); --slate-border: rgba(100,116,139,0.22);
  --shadow-sm: 0 1px 4px rgba(0,0,0,0.40);
  --shadow-md: 0 4px 24px rgba(0,0,0,0.52);
  --shadow-lg: 0 16px 60px rgba(0,0,0,0.65);
  --shadow-glow: 0 0 0 1px rgba(59,130,246,0.09), 0 8px 40px rgba(59,130,246,0.18);
  --pattern: rgba(59,130,246,0.014);
  --nav-bg: rgba(3,7,18,0.97);
  --nav-border: rgba(37,99,235,0.20);
  --input-bg: rgba(7,16,40,0.85);
  --toggle-bg: rgba(255,255,255,0.06);
  --card-stripe-blue: linear-gradient(90deg,#3b82f6,#22d3ee);
  --card-stripe-red: linear-gradient(90deg,#ef4444,#f87171);
  --card-stripe-amber: linear-gradient(90deg,#f59e0b,#fbbf24);
  --card-stripe-green: linear-gradient(90deg,#10b981,#34d399);
"""

_MAIN_CSS = (
    """
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/tailwindcss/2.2.19/tailwind.min.css">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;600;700&display=swap" rel="stylesheet">
<link href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@24,400,0,0" rel="stylesheet">
<style>
:root {"""
    + _VARS_LIGHT
    + """}
@media (prefers-color-scheme: dark) { :root {"""
    + _VARS_DARK
    + """} }

/* ── Reset & Base ─────────────────────────────────────────────────── */
*, *::before, *::after { box-sizing: border-box !important; }
html, body, [class*="st-"], .stApp {
  font-family: 'Inter', -apple-system, BlinkMacSystemFont, system-ui, sans-serif !important;
  -webkit-font-smoothing: antialiased !important; -moz-osx-font-smoothing: grayscale !important;
}
.stApp {
  background: var(--bg-base) !important;
  background-image: radial-gradient(circle at 1px 1px, var(--pattern) 1px, transparent 0) !important;
  background-size: 24px 24px !important; min-height: 100vh !important;
  transition: background 0.4s ease !important;
}
#MainMenu, footer, header { visibility: hidden !important; }
[data-testid="stSidebar"] { display: none !important; }
.block-container { padding: 0 2rem 6rem !important; max-width: 1440px !important; margin: 0 auto !important; }

/* ── Navbar ───────────────────────────────────────────────────────── */
.cs-nav {
  position: sticky; top: 0; z-index: 1000; overflow: hidden;
  display: flex; align-items: center; justify-content: space-between; gap: 16px;
  padding: 0 24px; height: 58px;
  background: var(--nav-bg);
  border-bottom: 1px solid var(--nav-border);
  backdrop-filter: blur(28px) saturate(200%); -webkit-backdrop-filter: blur(28px) saturate(200%);
  margin: 0 -2rem 1.5rem;
  animation: cs-navbar-in 0.55s cubic-bezier(0.16,1,0.3,1) both;
}
.cs-nav::after {
  content: ''; position: absolute; bottom: 0; left: -100%; width: 80px; height: 1px;
  background: linear-gradient(90deg, transparent, var(--accent), transparent);
  animation: cs-nav-scan 5s ease-in-out infinite;
}
.cs-nav-left { display: flex; align-items: center; gap: 11px; }
.cs-nav-logo-ring {
  width: 34px; height: 34px; flex-shrink: 0;
  background: linear-gradient(145deg, #3062e8, #1540c4);
  border-radius: 9px; display: flex; align-items: center; justify-content: center;
  box-shadow: 0 2px 12px rgba(37,99,235,0.45), inset 0 1px 0 rgba(255,255,255,0.15);
}
.cs-nav-brand { font-size: 13px; font-weight: 800; letter-spacing: -0.04em; color: var(--text-primary); line-height: 1; }
.cs-nav-tagline { font-size: 9px; font-weight: 600; letter-spacing: 0.14em; text-transform: uppercase; color: var(--text-muted); margin-top: 2px; }
.cs-nav-right { display: flex; align-items: center; gap: 8px; }
.cs-nav-user { font-size: 11px; font-weight: 700; letter-spacing: 0.04em; color: var(--text-secondary); padding: 4px 10px; background: var(--toggle-bg); border-radius: 20px; }
.cs-nav-ts { font-size: 10px; font-family: 'JetBrains Mono', monospace; color: var(--text-muted); letter-spacing: 0.03em; }
.cs-live-dot { width: 7px; height: 7px; border-radius: 50%; background: var(--green); display: inline-block; box-shadow: 0 0 8px var(--green); animation: cs-live 2s ease-in-out infinite; }
.cs-nav-divider { color: var(--border-strong); font-size: 16px; font-weight: 200; }

/* ── Metric Cards ─────────────────────────────────────────────────── */
.cs-bento {
  display: grid; grid-template-columns: repeat(4,1fr); gap: 14px; margin: 0 0 20px;
  animation: cs-rise 0.6s cubic-bezier(0.16,1,0.3,1) 0.1s both;
}
.cs-card {
  position: relative; overflow: hidden;
  background: var(--bg-surface);
  backdrop-filter: blur(24px) saturate(160%); -webkit-backdrop-filter: blur(24px) saturate(160%);
  border: 1px solid var(--border-subtle); border-radius: 20px; padding: 20px 22px 18px;
  box-shadow: var(--shadow-md);
  transition: transform 0.3s cubic-bezier(0.16,1,0.3,1), box-shadow 0.3s ease, border-color 0.25s;
}
.cs-card:hover { transform: translateY(-5px); box-shadow: var(--shadow-lg); border-color: var(--border-strong); }
.cs-card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px; }
.cs-card-blue::before { background: var(--card-stripe-blue); }
.cs-card-red::before { background: var(--card-stripe-red); }
.cs-card-amber::before { background: var(--card-stripe-amber); }
.cs-card-green::before { background: var(--card-stripe-green); }
.cs-card-icon { font-size: 18px; width: 38px; height: 38px; display: flex; align-items: center; justify-content: center; border-radius: 10px; margin-bottom: 14px; }
.cs-card-blue .cs-card-icon { background: var(--accent-light); }
.cs-card-red .cs-card-icon { background: var(--red-bg); }
.cs-card-amber .cs-card-icon { background: var(--amber-bg); }
.cs-card-green .cs-card-icon { background: var(--green-bg); }
.cs-card-value { font-size: 38px; font-weight: 700; letter-spacing: -0.05em; font-family: 'JetBrains Mono', monospace; line-height: 1; }
.cs-card-blue .cs-card-value { color: var(--accent); text-shadow: 0 0 24px var(--accent-glow); }
.cs-card-red .cs-card-value { color: var(--red); text-shadow: 0 0 18px rgba(220,38,38,0.22); }
.cs-card-amber .cs-card-value { color: var(--amber); text-shadow: 0 0 18px rgba(217,119,6,0.22); }
.cs-card-green .cs-card-value { color: var(--green); text-shadow: 0 0 18px rgba(5,150,105,0.22); }
.cs-card-label { font-size: 10px; font-weight: 700; letter-spacing: 0.10em; text-transform: uppercase; color: var(--text-muted); margin-top: 8px; }
.cs-card-sub { font-size: 10px; color: var(--text-muted); margin-top: 3px; opacity: 0.75; }

/* ── Chart Panels ─────────────────────────────────────────────────── */
.cs-charts-row { display: grid; grid-template-columns: 1fr 1.6fr; gap: 14px; margin-bottom: 20px; animation: cs-rise 0.6s cubic-bezier(0.16,1,0.3,1) 0.18s both; }
.cs-chart-card { background: var(--bg-surface); border: 1px solid var(--border-subtle); border-radius: 20px; overflow: hidden; box-shadow: var(--shadow-sm); padding: 8px 4px 4px; }
.cs-chart-title { font-size: 10px; font-weight: 700; letter-spacing: 0.12em; text-transform: uppercase; color: var(--text-muted); padding: 8px 20px 0; }
.cs-ok-panel { display: flex; align-items: center; justify-content: center; min-height: 200px; border-radius: 20px; background: var(--green-bg); border: 1px solid var(--green-border); font-size: 14px; font-weight: 700; color: var(--green); text-align: center; flex-direction: column; gap: 8px; }
.cs-ok-panel span { font-size: 32px; }

/* ── Health Bar ──────────────────────────────────────────────────── */
.cs-health-bar-wrap { background: var(--bg-surface); border: 1px solid var(--border-subtle); border-radius: 16px; padding: 14px 20px; margin-bottom: 20px; display: flex; align-items: center; gap: 16px; animation: cs-rise 0.6s cubic-bezier(0.16,1,0.3,1) 0.25s both; }
.cs-health-label { font-size: 10px; font-weight: 700; letter-spacing: 0.12em; text-transform: uppercase; color: var(--text-muted); white-space: nowrap; }
.cs-health-track { flex: 1; height: 6px; background: var(--border-subtle); border-radius: 3px; overflow: hidden; }
.cs-health-fill { height: 100%; border-radius: 3px; transition: width 0.8s cubic-bezier(0.16,1,0.3,1); }
.cs-health-pct { font-size: 13px; font-weight: 700; font-family: 'JetBrains Mono', monospace; white-space: nowrap; }

/* ── Section Headers ─────────────────────────────────────────────── */
.cs-section { font-size: 10px; font-weight: 700; letter-spacing: 0.14em; text-transform: uppercase; color: var(--text-muted); display: flex; align-items: center; gap: 8px; margin: 22px 0 12px; }
.cs-section::after { content: ''; flex: 1; height: 1px; background: var(--border-subtle); }

/* ── Filter Panel ────────────────────────────────────────────────── */
.cs-panel { background: var(--bg-surface); backdrop-filter: blur(16px); -webkit-backdrop-filter: blur(16px); border: 1px solid var(--border-subtle); border-radius: 20px; padding: 20px 24px; box-shadow: var(--shadow-sm); margin-bottom: 20px; animation: cs-rise 0.5s cubic-bezier(0.16,1,0.3,1) 0.2s both; }

/* ── Login ───────────────────────────────────────────────────────── */
.cs-login-wrap { padding: max(20px, calc(50vh - 240px)) 20px 0; display: flex; justify-content: center; }
.cs-login-card {
  width: 100%; max-width: 395px; position: relative;
  background: var(--bg-surface); backdrop-filter: blur(40px) saturate(200%); -webkit-backdrop-filter: blur(40px) saturate(200%);
  border: 1px solid var(--border-subtle); border-radius: 24px; padding: 44px 36px 300px;
  box-shadow: var(--shadow-lg), var(--shadow-glow);
  animation: cs-scale-in 0.55s cubic-bezier(0.16,1,0.3,1) both;
}
.cs-login-card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px; background: var(--card-stripe-blue); }
.cs-login-logo-ring { width: 58px; height: 58px; background: linear-gradient(145deg,#3062e8,#1540c4); border-radius: 16px; display: flex; align-items: center; justify-content: center; box-shadow: 0 4px 24px rgba(37,99,235,0.42), inset 0 1px 0 rgba(255,255,255,0.15); margin: 0 auto 16px; }
.cs-login-company { font-size: 20px; font-weight: 800; letter-spacing: -0.04em; color: var(--text-primary); text-align: center; }
.cs-login-subtitle { font-size: 11px; color: var(--text-muted); font-weight: 500; text-align: center; margin-bottom: 30px; letter-spacing: 0.06em; text-transform: uppercase; }
.cs-badge { display: inline-flex; align-items: center; gap: 6px; font-size: 10px; font-weight: 700; letter-spacing: 0.06em; text-transform: uppercase; color: var(--accent); background: var(--accent-light); border: 1px solid rgba(37,99,235,0.18); padding: 4px 10px; border-radius: 20px; }

/* ── Buttons ─────────────────────────────────────────────────────── */
.stButton > button {
  background: var(--accent) !important; color: #fff !important; border: none !important;
  border-radius: 10px !important; font-family: 'Inter', sans-serif !important;
  font-weight: 600 !important; font-size: 13px !important; letter-spacing: -0.01em !important;
  padding: 10px 18px !important; box-shadow: 0 2px 10px var(--accent-glow) !important;
  transition: all 0.2s cubic-bezier(0.16,1,0.3,1) !important; width: 100% !important;
}
.stButton > button:hover { background: var(--accent-hover) !important; transform: translateY(-1px) !important; box-shadow: 0 6px 20px var(--accent-glow) !important; }
.stButton > button:active { transform: translateY(0) !important; }
.cs-btn-ghost .stButton > button { background: var(--toggle-bg) !important; color: var(--text-secondary) !important; border: 1px solid var(--border-subtle) !important; box-shadow: none !important; }
.cs-btn-ghost .stButton > button:hover { background: var(--border-subtle) !important; color: var(--text-primary) !important; transform: none !important; box-shadow: none !important; }
.cs-btn-danger .stButton > button { background: var(--red-bg) !important; color: var(--red) !important; border: 1px solid var(--red-border) !important; box-shadow: none !important; }

/* ── Inputs ──────────────────────────────────────────────────────── */
.stTextInput > div > div > input, .stSelectbox > div > div, .stMultiselect > div > div { background: var(--input-bg) !important; border: 1px solid var(--border-strong) !important; border-radius: 10px !important; color: var(--text-primary) !important; font-family: 'Inter', sans-serif !important; font-size: 13px !important; transition: border-color 0.2s, box-shadow 0.2s !important; }
.stTextInput > div > div > input:focus { border-color: var(--accent) !important; box-shadow: 0 0 0 3px var(--accent-glow) !important; }
label[data-baseweb="form-control-label"] { color: var(--text-secondary) !important; font-size: 12px !important; font-weight: 600 !important; font-family: 'Inter', sans-serif !important; }
[data-baseweb="tag"] { background: var(--accent-light) !important; color: var(--accent) !important; border: 1px solid rgba(37,99,235,0.15) !important; border-radius: 6px !important; }
.stTextInput > div > div > input::placeholder { color: var(--placeholder-color) !important; opacity: 0.78 !important; }

/* ── Data Table ──────────────────────────────────────────────────── */
.stDataFrame { border-radius: 16px !important; overflow: hidden !important; }
.stDataFrame [data-testid="stDataFrameResizable"] { border: 1px solid var(--border-subtle) !important; border-radius: 16px !important; }

/* ── Plotly ──────────────────────────────────────────────────────── */
.stPlotlyChart { border-radius: 16px !important; overflow: hidden !important; background: transparent !important; }

/* ── Expander ────────────────────────────────────────────────────── */
[data-testid="stExpander"] {
  border: 1px solid var(--border-subtle) !important;
  border-radius: 12px !important;
  overflow: hidden !important;
}
[data-testid="stExpander"] > details > summary {
  background: var(--bg-surface) !important;
  border: none !important;
  border-radius: 0 !important;
  color: var(--text-secondary) !important;
  font-family: 'Inter', sans-serif !important;
  font-size: 12px !important;
  font-weight: 600 !important;
  list-style: none !important;
  padding: 10px 14px 10px 34px !important;
  position: relative !important;
}
[data-testid="stExpander"] > details > summary::-webkit-details-marker {
  display: none !important;
}
[data-testid="stExpander"] > details[open] > summary {
  border-bottom: 1px solid var(--border-subtle) !important;
}
[data-testid="stExpander"] > details > div {
  background: var(--bg-glass) !important;
  border: none !important;
  border-radius: 0 !important;
}
/* Reemplaza el icono Material para evitar texto tipo _arrow_ si falla la fuente */
[data-testid="stExpander"] summary .material-symbols-outlined,
[data-testid="stExpander"] summary span[class*="material-symbols"],
[data-testid="stExpander"] summary [data-testid="stExpanderToggleIcon"] {
  display: none !important;
  font-size: 0 !important;
  width: 0 !important;
  overflow: hidden !important;
}
[data-testid="stExpander"] > details > summary::before {
  content: "›";
  position: absolute;
  left: 14px;
  top: 50%;
  transform: translateY(-50%);
  color: var(--text-secondary);
  font-family: Arial, sans-serif;
  font-size: 20px;
  font-weight: 700;
  line-height: 1;
  transition: transform 0.15s ease;
}
[data-testid="stExpander"] > details[open] > summary::before {
  transform: translateY(-50%) rotate(90deg);
}

/* ── Alerts ──────────────────────────────────────────────────────── */
.stSuccess, .stInfo, .stWarning, .stError { border-radius: 12px !important; font-size: 13px !important; font-family: 'Inter', sans-serif !important; }
.stCaption { color: var(--text-secondary) !important; font-size: 12px !important; font-family: 'Inter', sans-serif !important; opacity: 0.95 !important; }

/* Texto auxiliar bajo exportar (mejor contraste en oscuro) */
.cs-export-stats {
  margin-top: 8px;
  font-size: 12px;
  font-weight: 600;
  color: var(--text-secondary);
  letter-spacing: 0.02em;
}

/* ── Animations ──────────────────────────────────────────────────── */
@keyframes cs-rise { from { opacity:0; transform:translateY(20px); } to { opacity:1; transform:translateY(0); } }
@keyframes cs-navbar-in { from { opacity:0; transform:translateY(-16px); } to { opacity:1; transform:translateY(0); } }
@keyframes cs-scale-in { from { opacity:0; transform:scale(0.95) translateY(10px); } to { opacity:1; transform:scale(1) translateY(0); } }
@keyframes cs-live { 0%,100% { box-shadow:0 0 6px var(--green); opacity:1; } 50% { box-shadow:0 0 14px var(--green); opacity:0.5; } }
@keyframes cs-nav-scan { 0% { left:-100%; } 100% { left:250%; } }
@keyframes cs-orb-drift { 0%,100% { transform:translate(0,0) scale(1); } 33% { transform:translate(35px,-25px) scale(1.04); } 66% { transform:translate(-18px,32px) scale(0.96); } }

/* ── Responsive ───────────────────────────────────────────────────── */
@media (max-width: 1100px) {
    .cs-bento { grid-template-columns: repeat(2, minmax(0,1fr)); }
    .cs-charts-row { grid-template-columns: 1fr; }
}
@media (max-width: 780px) {
    .block-container { padding: 0 1rem 2.5rem !important; }
    .cs-nav { height: auto; padding: 10px 14px; gap: 10px; flex-wrap: wrap; margin: 0 -1rem 1rem; }
    .cs-nav-right { width: 100%; justify-content: space-between; }
    .cs-bento { grid-template-columns: 1fr; gap: 10px; }
    .cs-card { border-radius: 16px; padding: 16px; }
    .cs-card-value { font-size: 30px; }
    .cs-panel { border-radius: 16px; padding: 14px; }
    .cs-health-bar-wrap { flex-direction: column; align-items: flex-start; gap: 8px; padding: 12px 14px; }
    .cs-health-track { width: 100%; }
}

@media print { .cs-nav, .stButton, .stDownloadButton { display: none !important; } }
</style>
"""
)


def _load_logo_data_uri() -> str:
    """Carga el logo PNG real y lo expone como data URI para HTML."""
    candidate_paths = [
        Path(__file__).resolve().parent / "ColsaborSAS.PNG",
        Path(__file__).resolve().parents[1] / "ColsaborSAS.PNG",
    ]
    for logo_path in candidate_paths:
        if logo_path.exists():
            try:
                return "data:image/png;base64," + base64.b64encode(
                    logo_path.read_bytes()
                ).decode("ascii")
            except Exception:  # pragma: no cover
                return ""
    return ""  # pragma: no cover


_LOGO_DATA_URI = _load_logo_data_uri()

_LOGO_SM = (
    f'<img src="{_LOGO_DATA_URI}" alt="Colsabor" '
    'style="width:20px;height:20px;object-fit:cover;object-position:center 72%;'
    'display:block;filter:brightness(1.25) contrast(1.15) saturate(1.2);"/>'
    if _LOGO_DATA_URI
    else (
        '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" xmlns="http://www.w3.org/2000/svg">'
        '<rect width="20" height="20" rx="6" fill="rgba(255,255,255,0.10)"/>'
        '<path d="M7 7.5C7 5.8 8.2 4.8 10 4.8C11.8 4.8 13 5.8 13 7.2C13 8.8 11.2 9.7 10 10.2'
        'C8.8 10.7 7 11.5 7 13C7 14.4 8.2 15.2 10 15.2C11.8 15.2 13 14.2 13 13.4"'
        ' stroke="white" stroke-width="1.6" stroke-linecap="round" fill="none"/></svg>'
    )
)

_LOGO_LG = (
    f'<div style="width:220px;max-width:100%;padding:18px 20px;'
    f"background:rgba(255,255,255,0.10);backdrop-filter:blur(12px);"
    f"border-radius:20px;border:1px solid rgba(255,255,255,0.15);"
    f'display:flex;justify-content:center;align-items:center;margin:0 auto 4px;">'
    f'<img src="{_LOGO_DATA_URI}" alt="Colsabor" '
    'style="width:180px;max-width:100%;height:auto;display:block;'
    "filter:brightness(1.15) contrast(1.08) saturate(1.1) "
    'drop-shadow(0 4px 12px rgba(0,0,0,0.35));"/>'
    "</div>"
    if _LOGO_DATA_URI
    else (
        '<svg width="120" height="48" viewBox="0 0 120 48" fill="none" xmlns="http://www.w3.org/2000/svg">'
        '<rect width="120" height="48" rx="12" fill="rgba(255,255,255,0.07)"/>'
        '<rect x="10" y="10" width="28" height="28" rx="8" fill="url(#lg)"/>'
        '<defs><linearGradient id="lg" x1="10" y1="10" x2="38" y2="38" gradientUnits="userSpaceOnUse">'
        '<stop stop-color="#3b82f6"/><stop offset="1" stop-color="#0891b2"/></linearGradient></defs>'
        '<text x="26" y="29" font-family="Arial,sans-serif" font-size="13" font-weight="900" '
        'fill="white" text-anchor="middle" letter-spacing="-0.5">CS</text>'
        '<text x="50" y="26" font-family="Arial,sans-serif" font-size="13" font-weight="800" '
        'fill="white" letter-spacing="-0.3">COLSABOR</text>'
        '<text x="50" y="38" font-family="Arial,sans-serif" font-size="7.5" font-weight="500" '
        'fill="rgba(255,255,255,0.55)" letter-spacing="1.2">S.A.S</text>'
        "</svg>"
    )
)


def _inject_css():
    theme = st.session_state.get("theme_override", "auto")
    st.html(_MAIN_CSS)
    if theme == "dark":
        st.html(f"<style>:root {{ {_VARS_DARK} }}</style>")
    elif theme == "light":
        st.html(
            f"<style>:root {{ {_VARS_LIGHT} }} @media (prefers-color-scheme: dark) {{ :root {{ {_VARS_LIGHT} }} }}</style>",
        )


def _render_loading(msg: str):
    """Pantalla de carga fullscreen estática con saludo, mensaje random y barra de progreso."""
    usuario_email = st.session_state.get("usuario_email", "")
    saludo = get_greeting(usuario_email)
    mensaje_random = get_random_message(usuario_email)

    # Obtener progreso de carga
    progreso_pct = st.session_state.get("_loading_progress", 0)
    progreso_text = st.session_state.get("_loading_text", "Inicializando...")

    # Detectar tema actual para aplicar colores correctos
    theme = st.session_state.get("theme_override", "auto")
    # Si es auto, usar variables de Streamlit para detectar
    # Inyectar variables de tema adecuadas
    theme_vars = _VARS_DARK if theme == "dark" else _VARS_LIGHT

    st.html(
        f"""
<style>
/* ── Variables de tema ──────────────────────────────────── */
:root {{
  {theme_vars}
}}

/* ── Bloquear scroll de toda la app ─────────────────────── */
html, body, [data-testid="stAppViewContainer"],
[data-testid="stMain"], .main, .block-container {{
  overflow: hidden !important;
  height: 100vh !important;
  max-height: 100vh !important;
}}

@keyframes cs-spin {{
  0%   {{ transform: rotate(0deg); }}
  100% {{ transform: rotate(360deg); }}
}}
@keyframes cs-bar-wave {{
  0%,100% {{ height: 8px;  opacity: .35; }}
  50%      {{ height: 28px; opacity: 1;   }}
}}
@keyframes cs-fade-in {{
  from {{ opacity: 0; transform: translateY(18px); }}
  to   {{ opacity: 1; transform: translateY(0); }}
}}
@keyframes cs-orb {{
  0%,100% {{ transform: translate(0,0) scale(1); }}
  50%     {{ transform: translate(20px,-12px) scale(1.06); }}
}}
@keyframes cs-progress-pulse {{
  0%,100% {{ box-shadow: 0 0 0 0 rgba(59,130,246,0.4); }}
  50%     {{ box-shadow: 0 0 0 8px rgba(59,130,246,0); }}
}}

/* ── Overlay fullscreen ──────────────────────────────────── */
.cs-splash {{
  position: fixed;
  inset: 0;
  z-index: 99999;
  display: flex;
  flex-direction: column;
  align-items: center;
  justify-content: center;
  background: var(--bg-base);
  overflow: hidden;
  animation: cs-fade-in .4s ease both;
}}

/* Orbs decorativos de fondo */
.cs-splash-orb {{
  position: absolute;
  border-radius: 50%;
  filter: blur(80px);
  pointer-events: none;
}}
.cs-splash-orb-1 {{
  width: 340px; height: 340px;
  background: radial-gradient(circle, rgba(37,99,235,.18) 0%, transparent 70%);
  top: -60px; left: -80px;
  animation: cs-orb 7s ease-in-out infinite;
}}
.cs-splash-orb-2 {{
  width: 280px; height: 280px;
  background: radial-gradient(circle, rgba(34,211,238,.14) 0%, transparent 70%);
  bottom: -50px; right: -60px;
  animation: cs-orb 9s ease-in-out infinite reverse;
}}

/* Logo mini en la esquina */
.cs-splash-logo {{
  position: absolute;
  top: 28px; left: 36px;
  display: flex; align-items: center; gap: 12px;
}}
.cs-splash-brand {{
  font-family: 'Inter', sans-serif;
  font-size: 13px; font-weight: 800;
  letter-spacing: .12em;
  color: var(--text-primary);
  opacity: .8;
}}
}}

/* Contenido central */
.cs-splash-body {{
  display: flex;
  flex-direction: column;
  align-items: center;
  gap: 0;
  text-align: center;
  padding: 0 32px;
  max-width: 560px;
  width: 100%;
  animation: cs-fade-in .6s .1s ease both;
}}

/* Saludo */
.cs-splash-greeting {{
  font-family: 'Inter', sans-serif;
  font-size: clamp(18px, 3vw, 26px);
  font-weight: 700;
  color: var(--text-primary);
  letter-spacing: -.03em;
  margin-bottom: 6px;
}}

/* Acción en progreso */
.cs-splash-action {{
  font-family: 'JetBrains Mono', monospace;
  font-size: 11px;
  letter-spacing: .12em;
  text-transform: uppercase;
  color: var(--amber);
  margin-bottom: 36px;
  opacity: .85;
}}

/* Anillo de carga */
.cs-splash-ring {{
  width: 76px; height: 76px;
  border-radius: 50%;
  border: 3px solid rgba(59,130,246,.12);
  border-top:   3px solid #3b82f6;
  border-right: 3px solid #22d3ee;
  animation: cs-spin 1.1s linear infinite;
  box-shadow: 0 0 36px rgba(59,130,246,.22);
  margin-bottom: 32px;
}}

/* Mensaje random */
.cs-splash-quote {{
  font-family: 'Inter', sans-serif;
  font-size: clamp(12px, 1.6vw, 14px);
  color: var(--text-muted);
  line-height: 1.6;
  max-width: 400px;
  margin-bottom: 36px;
  font-style: italic;
}}

/* Barras ecualizador */
.cs-splash-bars {{
  display: flex;
  align-items: flex-end;
  gap: 5px;
  height: 32px;
  margin-bottom: 28px;
}}
.cs-splash-bar {{
  width: 5px;
  border-radius: 3px;
  background: linear-gradient(180deg, #3b82f6, #22d3ee);
  animation: cs-bar-wave 1.3s ease-in-out infinite;
}}
.cs-splash-bar:nth-child(1) {{ animation-delay: 0s;    height: 10px; }}
.cs-splash-bar:nth-child(2) {{ animation-delay: .18s;  height: 18px; }}
.cs-splash-bar:nth-child(3) {{ animation-delay: .36s;  height: 28px; }}
.cs-splash-bar:nth-child(4) {{ animation-delay: .54s;  height: 18px; }}
.cs-splash-bar:nth-child(5) {{ animation-delay: .72s;  height: 10px; }}

/* Barra de progreso */
.cs-progress-container {{
  width: 100%;
  max-width: 340px;
  display: flex;
  flex-direction: column;
  gap: 8px;
  margin-bottom: 24px;
}}
.cs-progress-track {{
  width: 100%;
  height: 6px;
  background: rgba(59,130,246,0.12);
  border-radius: 3px;
  overflow: hidden;
  border: 1px solid rgba(59,130,246,0.2);
}}
.cs-progress-fill {{
  height: 100%;
  background: linear-gradient(90deg, #3b82f6, #22d3ee);
  width: {progreso_pct}%;
  border-radius: 3px;
  transition: width 0.3s ease;
  box-shadow: 0 0 12px rgba(59,130,246,0.4);
}}
.cs-progress-text {{
  display: flex;
  justify-content: space-between;
  align-items: center;
  font-family: 'JetBrains Mono', monospace;
  font-size: 10px;
  color: var(--text-muted);
  letter-spacing: .05em;
}}
.cs-progress-percent {{
  color: #3b82f6;
  font-weight: 700;
}}
.cs-progress-label {{
  max-width: 200px;
  overflow: hidden;
  text-overflow: ellipsis;
  white-space: nowrap;
  text-align: right;
}}

/* Tag estado */
.cs-splash-status {{
  position: absolute;
  bottom: 32px;
  font-family: 'JetBrains Mono', monospace;
  font-size: 10px;
  letter-spacing: .1em;
  color: var(--text-muted, #64748b);
  text-transform: uppercase;
  opacity: .6;
}}
</style>

<div class="cs-splash">
  <!-- orbs de fondo -->
  <div class="cs-splash-orb cs-splash-orb-1"></div>
  <div class="cs-splash-orb cs-splash-orb-2"></div>

  <!-- Logo esquina -->
  <div class="cs-splash-logo">
    <div class="cs-splash-brand">COLSABOR</div>
  </div>

  <!-- Centro -->
  <div class="cs-splash-body">
    <div class="cs-splash-greeting">{saludo}</div>
    <div class="cs-splash-action">{msg}</div>
    <div class="cs-splash-ring"></div>
    <div class="cs-splash-quote">"{mensaje_random}"</div>
    <div class="cs-splash-bars">
      <div class="cs-splash-bar"></div>
      <div class="cs-splash-bar"></div>
      <div class="cs-splash-bar"></div>
      <div class="cs-splash-bar"></div>
      <div class="cs-splash-bar"></div>
    </div>
    
    <!-- Barra de progreso -->
    <div class="cs-progress-container">
      <div class="cs-progress-track">
        <div class="cs-progress-fill"></div>
      </div>
      <div class="cs-progress-text">
        <span class="cs-progress-label">{progreso_text}</span>
        <span class="cs-progress-percent">{progreso_pct}%</span>
      </div>
    </div>
  </div>

  <!-- Pie de página -->
  <div class="cs-splash-status">Monitor de Inventario · Colsabor S.A.S</div>
</div>
"""
    )


# ── Plotly chart helpers ──────────────────────────────────────────────────────


def _build_donut_chart(
    criticos: int, bajos: int, ok: int, no_encontrados: int, total: int
):
    """Donut chart showing inventory status distribution."""
    labels = ["Crítico", "Bajo", "OK", "No encontrado"]
    values = [criticos, bajos, ok, no_encontrados]
    colors = ["#ef4444", "#f59e0b", "#10b981", "#64748b"]
    pct_ok = round(100 * ok / total) if total else 0

    fig = go.Figure(
        go.Pie(
            labels=labels,
            values=values,
            hole=0.70,
            marker=dict(colors=colors, line=dict(width=0)),
            textinfo="none",
            hovertemplate="%{label}: %{value}<br>%{percent}<extra></extra>",
        )
    )
    fig.add_annotation(
        text=(
            f"<b style='font-size:22px'>{pct_ok}%</b>"
            "<br><span style='font-size:9px;color:#64748b;letter-spacing:2px'>SALUDABLE</span>"
        ),
        x=0.5,
        y=0.5,
        showarrow=False,
        align="center",
        font=dict(size=14, family="JetBrains Mono, monospace", color="#94a3b8"),
    )
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(t=10, b=10, l=10, r=10),
        height=240,
        showlegend=True,
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=-0.28,
            xanchor="center",
            x=0.5,
            font=dict(size=10, color="#64748b", family="Inter, sans-serif"),
        ),
        font=dict(family="Inter, sans-serif"),
    )
    return fig


def _build_deficit_chart(df_deficit: pd.DataFrame):
    """Horizontal bar chart of top items with the largest stock deficit."""
    labels = (
        df_deficit["Referencia"].str[:10] + " · " + df_deficit["Nombre"].str[:16]
    ).tolist()
    values = df_deficit["Diferencia"].tolist()
    bar_colors = [
        "#ef4444" if "Crítico" in str(e) else "#f59e0b"
        for e in df_deficit["Estado"].tolist()
    ]
    fig = go.Figure(
        go.Bar(
            x=values,
            y=labels,
            orientation="h",
            marker=dict(color=bar_colors, opacity=0.88, line=dict(width=0)),
            text=[f"{v:,.0f}g" for v in values],
            textposition="outside",
            textfont=dict(size=10, color="#94a3b8", family="JetBrains Mono, monospace"),
            hovertemplate="%{y}<br><b>Déficit: %{x:,.0f} g</b><extra></extra>",
        )
    )
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(t=10, b=24, l=8, r=88),
        height=240,
        xaxis=dict(
            showgrid=True,
            gridcolor="rgba(100,116,139,0.10)",
            gridwidth=1,
            color="#475569",
            zeroline=True,
            zerolinecolor="rgba(100,116,139,0.20)",
            tickfont=dict(size=9, family="JetBrains Mono, monospace"),
        ),
        yaxis=dict(
            color="#94a3b8",
            showgrid=False,
            automargin=True,
            tickfont=dict(size=10, family="Inter, sans-serif"),
        ),
        font=dict(family="Inter, sans-serif", color="#94a3b8"),
        bargap=0.32,
    )
    return fig


def main():
    """Función principal de la aplicación."""
    _inject_css()

    # ── LOGIN ────────────────────────────────────────────────────────────────
    if "usuario_email" not in st.session_state:
        _theme = st.session_state.get("theme_override", "auto")
        _theme_icon = "☀️" if _theme == "dark" else "🌙"

        # Cosmic Forge solo en oscuro/auto; modo claro usa el fondo de _MAIN_CSS
        if _theme != "light":
            st.html(
                """<style>
.stApp {
  background: #010814 !important;
  background-image:
    radial-gradient(ellipse 80% 70% at 10% 20%, rgba(59,130,246,0.24) 0%, transparent 60%),
    radial-gradient(ellipse 65% 80% at 90% 85%, rgba(6,182,212,0.20) 0%, transparent 60%),
    radial-gradient(ellipse 55% 50% at 58% 0%, rgba(99,102,241,0.16) 0%, transparent 60%),
    radial-gradient(circle at 1px 1px, rgba(59,130,246,0.07) 1px, transparent 0) !important;
  background-size: auto, auto, auto, 28px 28px !important;
}
</style>"""
            )

        # CSS del login: centrado vertical via flex, card sobre st.container y
        # toggle fijo arriba-derecha. st.html() evita que Streamlit elimine
        # los <style> tags (a diferencia de st.markdown()).
        _, col, _ = st.columns([1, 1.1, 1])
        with col:
            st.html(
                "<style>"
                "html,body{overflow:hidden!important}"
                ".block-container{display:flex!important;align-items:center"
                "!important;min-height:100vh!important;padding:0 1.5rem"
                "!important;max-width:100%!important;margin:0!important}"
                ".block-container>div[data-testid='stVerticalBlock']"
                "{width:100%}"
                ".st-key-login_card{background:var(--bg-surface)!important;"
                "backdrop-filter:blur(40px) saturate(200%);"
                "-webkit-backdrop-filter:blur(40px) saturate(200%);"
                "border:1px solid var(--border-subtle)!important;"
                "border-radius:24px!important;"
                "padding:40px 36px 36px!important;"
                "box-shadow:var(--shadow-lg),var(--shadow-glow)!important;"
                "position:relative!important;"
                "animation:cs-scale-in 0.55s cubic-bezier(0.16,1,0.3,1) both}"
                ".st-key-login_theme{position:fixed!important;top:16px"
                "!important;right:16px!important;z-index:1100!important;"
                "width:auto!important}"
                ".st-key-login_theme button{width:44px!important;"
                "min-width:44px!important;height:44px!important;"
                "border-radius:999px!important;padding:0!important;"
                "font-size:20px!important}"
                "</style>"
            )
            with st.container(key="login_card"):
                st.markdown(
                    f'<div style="display:flex;justify-content:center;'
                    f'margin-bottom:20px">{_LOGO_LG}</div>'
                    f'<div class="cs-login-subtitle">Sistema de Inventario'
                    f"</div>",
                    unsafe_allow_html=True,
                )
                st.markdown('<div class="cs-btn-ghost">', unsafe_allow_html=True)
                if st.button(_theme_icon, help="Cambiar tema", key="login_theme"):
                    st.session_state["theme_override"] = (
                        "light" if _theme == "dark" else "dark"
                    )
                    st.rerun()
                st.markdown("</div>", unsafe_allow_html=True)
                st.markdown(
                    '<div style="text-align:center;margin-bottom:20px">'
                    '<span class="cs-badge">🔒 Acceso Restringido</span></div>',
                    unsafe_allow_html=True,
                )
                usuario_email = st.text_input(
                    "Correo corporativo",
                    placeholder="nombre@colsabor.com.co",
                    key="email_input",
                )
                usuario_password = st.text_input(
                    "Contraseña",
                    type="password",
                    placeholder="••••••••",
                    key="password_input",
                )
                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                if st.button(
                    "Iniciar Sesión →", use_container_width=True, type="primary"
                ):
                    email_clean = usuario_email.strip().lower()
                    _log.debug(
                        "[LOGIN] intento -> email=%s pass_len=%d",
                        email_clean,
                        len(usuario_password),
                    )
                    # Reiniciar log de UI
                    st.session_state["_dbg_log"] = []

                    def _dbg(msg: str):
                        st.session_state.setdefault("_dbg_log", []).append(msg)
                        _log.debug("[LOGIN-UI] %s", msg)

                    if not email_clean or not usuario_password:
                        _dbg("⚠️ Correo o contraseña vacíos")
                        st.warning("Completa tu correo y contraseña.")
                    elif email_clean not in ALLOWED_EMAILS:
                        _dbg(f"❌ Email no autorizado: {email_clean}")
                        _log.debug("[LOGIN] email no autorizado: %s", email_clean)
                        st.error("Acceso denegado. Este correo no está autorizado.")
                    else:
                        _dbg(f"✅ Email autorizado: {email_clean}")
                        _log.debug(
                            "[LOGIN] email autorizado, obteniendo cliente Supabase..."
                        )
                        _dbg(
                            f"🔑 URL={SUPABASE_URL[:30] if SUPABASE_URL else 'VACÍO'} KEY_len={len(SUPABASE_KEY or '')}"
                        )
                        supabase = get_supabase_client()
                        if supabase is None:
                            _dbg("❌ get_supabase_client() devolvió None")
                            _log.debug("[LOGIN] supabase client es None -> fallo")
                            st.error("Sin conexión a Supabase.")
                        else:
                            _dbg("✅ Cliente Supabase creado OK")
                            _log.debug(
                                "[LOGIN] cliente OK, llamando sign_in_with_password..."
                            )
                            with st.spinner("Verificando credenciales…"):
                                try:
                                    _dbg("⏳ Llamando sign_in_with_password...")
                                    resp = supabase.auth.sign_in_with_password(
                                        {
                                            "email": email_clean,
                                            "password": usuario_password,
                                        }
                                    )
                                    _log.debug(
                                        "[LOGIN] respuesta -> session=%s user=%s",
                                        bool(resp.session),
                                        resp.user.email if resp.user else None,
                                    )
                                    if resp.session:
                                        _dbg("✅ LOGIN EXITOSO — redirigiendo...")
                                        st.session_state["usuario_email"] = (
                                            resp.user.email if resp.user else None
                                        ) or email_clean
                                        # Renueva token Siigo más adelante al cargar datos.
                                        if "token_siigo" in st.session_state:
                                            del st.session_state["token_siigo"]
                                        _log.debug(
                                            "[LOGIN] LOGIN EXITOSO para %s", email_clean
                                        )
                                        st.rerun()
                                    else:
                                        _dbg("❌ Respuesta sin session")
                                        _log.debug(
                                            "[LOGIN] sin session en la respuesta"
                                        )
                                        st.error("Credenciales incorrectas.")
                                except Exception as e:
                                    msg = str(e)
                                    _dbg(f"💥 Excepción: {msg}")
                                    _log.debug("[LOGIN] excepcion en sign_in: %s", msg)
                                    if "Invalid login credentials" in msg:
                                        st.error("Correo o contraseña incorrectos.")
                                    elif "Email not confirmed" in msg:
                                        st.warning(
                                            "Confirma tu correo antes de entrar."
                                        )
                                    else:
                                        st.error(f"Error: {msg}")

                # Mostrar log de último intento de login
                if st.session_state.get("_dbg_log"):
                    lines = "\n".join(st.session_state["_dbg_log"])
                    st.markdown(
                        f'<div style="margin-top:10px;padding:10px 14px;border-radius:10px;'
                        f"background:rgba(0,0,0,0.40);border:1px solid rgba(99,102,241,0.25);"
                        f"font-family:JetBrains Mono,monospace;font-size:11px;line-height:1.8;"
                        f'color:#c4b5fd;white-space:pre-wrap">'
                        f'<span style="color:#818cf8;font-size:9px;letter-spacing:.1em;font-weight:700">ÚLTIMO INTENTO</span>\n'
                        f"{lines}</div>",
                        unsafe_allow_html=True,
                    )

                # ── Panel de debug siempre visible ───────────────────────
                _url_ok = bool(SUPABASE_URL)
                _key_ok = (SUPABASE_KEY or "").startswith("eyJ")
                _secrets_ok = bool(st.secrets.get("SUPABASE_URL", ""))
                _status_url = "✅ Configurada" if _url_ok else "❌ No encontrada"
                _status_key = "✅ JWT válido" if _key_ok else "❌ Formato incorrecto"
                _status_src = (
                    "☁️ Streamlit Cloud" if _secrets_ok else "📦 Fallback interno"
                )
                # ── Panel de estado de dependencias ──────────────────────
                def _chk_import(pkg: str) -> str:
                    try:
                        __import__(pkg)
                        return "✅"
                    except ImportError:
                        return "❌"

                _deps = [
                    ("streamlit", "streamlit"),
                    ("pandas", "pandas"),
                    ("openpyxl", "openpyxl"),
                    ("requests", "requests"),
                    ("plotly", "plotly"),
                    ("supabase", "supabase"),
                ]
                _dep_rows = "".join(
                    f"<div>{_chk_import(pkg)}&nbsp;<span style='color:#cbd5e1'>{label}</span></div>"
                    for label, pkg in _deps
                )
                st.markdown(
                    f"""
<div style="margin-top:16px;padding:12px 14px;border-radius:12px;
background:rgba(0,0,0,0.35);border:1px solid rgba(255,255,255,0.10);
font-family:'JetBrains Mono',monospace;font-size:11px;line-height:1.9;
color:#94a3b8">
<div style="color:#60a5fa;font-weight:700;margin-bottom:6px;font-size:10px;
letter-spacing:.12em">🔍 ESTADO CONEXIÓN SUPABASE</div>
<div>URL &nbsp;&nbsp;&nbsp;&nbsp;→ <span style="color:#e2e8f0">{_status_url}</span></div>
<div>API KEY → <span style="color:#e2e8f0">{_status_key}</span></div>
<div>Fuente &nbsp;→ <span style="color:#e2e8f0">{_status_src}</span></div>
<div style="color:#60a5fa;font-weight:700;margin:10px 0 6px;font-size:10px;
letter-spacing:.12em">📦 DEPENDENCIAS INSTALADAS</div>
{_dep_rows}
</div>""",
                    unsafe_allow_html=True,
                )
        st.stop()

    # ── NAVBAR ───────────────────────────────────────────────────────────────
    usuario_email = st.session_state.get("usuario_email", "")
    ultima_act = st.session_state.get("ultima_actualizacion", now_colombia())
    theme = st.session_state.get("theme_override", "auto")
    theme_icon = "☀️" if theme == "dark" else "🌙"

    st.markdown(
        f"""
        <div class="cs-nav">
          <div class="cs-nav-left">
            <div class="cs-nav-logo-ring">{_LOGO_SM}</div>
            <div>
              <div class="cs-nav-brand">COLSABOR</div>
              <div class="cs-nav-tagline">Monitor de Inventario</div>
            </div>
          </div>
          <div class="cs-nav-right">
            <span class="cs-live-dot"></span>
            <span class="cs-nav-ts" style="color:var(--green);font-weight:700;font-size:10px;letter-spacing:.1em">LIVE</span>
            <span class="cs-nav-ts" style="color:var(--border-subtle)">|</span>
            <span class="cs-nav-ts">{now_colombia().strftime('%d %b %Y · %H:%M:%S')}</span>
            <span class="cs-nav-ts" style="color:var(--border-subtle)">|</span>
            <span class="cs-nav-ts" style="font-size:9px;opacity:.7">act. {ultima_act.strftime('%H:%M')}</span>
            <span class="cs-nav-user">{usuario_email.split('@')[0].upper()}</span>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    current_page = st.session_state.get("current_page", "monitor")

    nav1, nav2, _spacer, nc1, nc2, nc3 = st.columns([1.3, 1.9, 3.3, 1, 1, 1])
    with nav1:
        is_monitor = current_page == "monitor"
        if is_monitor:
            st.button("📦 Monitor", use_container_width=True, type="primary", key="nav_monitor")
        else:
            st.markdown('<div class="cs-btn-ghost">', unsafe_allow_html=True)
            if st.button("📦 Monitor", use_container_width=True, key="nav_monitor"):
                st.session_state["current_page"] = "monitor"
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
    with nav2:
        is_dane = current_page == "dane"
        if is_dane:
            st.button("📊 Encuesta DANE", use_container_width=True, type="primary", key="nav_dane")
        else:
            st.markdown('<div class="cs-btn-ghost">', unsafe_allow_html=True)
            if st.button("📊 Encuesta DANE", use_container_width=True, key="nav_dane"):
                st.session_state["current_page"] = "dane"
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
    with nc1:
        st.markdown('<div class="cs-btn-ghost">', unsafe_allow_html=True)
        if st.button(theme_icon, help="Cambiar tema", use_container_width=True):
            st.session_state["theme_override"] = "light" if theme == "dark" else "dark"
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)
    with nc2:
        st.markdown('<div class="cs-btn-ghost">', unsafe_allow_html=True)
        if st.button("⟳", help="Actualizar datos de Siigo", use_container_width=True):
            st.session_state["forzar_actualizacion"] = True
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)
    with nc3:
        st.markdown('<div class="cs-btn-danger">', unsafe_allow_html=True)
        if st.button("↩", help="Cerrar sesión", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    # ── ENRUTAMIENTO DE PÁGINAS ───────────────────────────────────────────────
    if current_page == "dane":
        dane_survey.render_dane_survey()
        # Footer
        st.markdown(
            """
            <div style="
              position:fixed;bottom:0;left:0;right:0;z-index:900;
              padding:10px 24px;
              background:var(--nav-bg);
              border-top:1px solid var(--nav-border);
              backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);
              display:flex;align-items:center;justify-content:center;
            ">
              <p style="color:var(--text-muted);font-size:11px;letter-spacing:0.06em;
                margin:0;font-family:'Inter',sans-serif;font-weight:500">
                COLSABOR S.A.S &nbsp;·&nbsp; Encuesta DANE · DIAN Automator &nbsp;·&nbsp; © 2026
              </p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.stop()

    # ── CARGA INVENTARIO MÍNIMO ───────────────────────────────────────────────
    needs_excel = (
        "df_excel_cache" not in st.session_state
        or "forzar_actualizacion" in st.session_state
    )
    needs_siigo = (
        "df_siigo_cache" not in st.session_state
        or "forzar_actualizacion" in st.session_state
    )

    if needs_excel or needs_siigo:
        # Mostrar loading screen antes de las llamadas de red
        _loading_start = time.time()

        # Inicializar progreso
        st.session_state["_loading_progress"] = 0
        st.session_state["_loading_text"] = "Inicializando..."
        _render_loading(
            "Descargando productos de Siigo…"
            if not needs_excel
            else "Cargando inventario mínimo…"
        )

        if needs_excel:
            st.session_state["_loading_progress"] = 10
            st.session_state["_loading_text"] = (
                "Cargando inventario mínimo desde Supabase..."
            )
            _render_loading("Cargando inventario mínimo…")

            df_excel = cargar_inventario_minimo_supabase()
            if df_excel is not None:
                st.session_state["df_excel_cache"] = df_excel
                st.session_state["_loading_progress"] = 35
                st.session_state["_loading_text"] = "Inventario mínimo cargado ✓"
            else:
                st.error("No se pudo cargar el inventario mínimo. Verifica Supabase.")
                st.stop()
        else:
            df_excel = st.session_state["df_excel_cache"]

        # ── CARGA SIIGO ───────────────────────────────────────────────────────
        if needs_siigo:
            datos_guardados = None
            if "forzar_actualizacion" not in st.session_state:
                st.session_state["_loading_progress"] = 40
                st.session_state["_loading_text"] = "Buscando datos en caché..."
                _render_loading("Buscando datos en caché…")
                datos_guardados = cargar_productos_siigo_guardados()

            if datos_guardados is not None:
                df_siigo, _, ultima_actualizacion = datos_guardados
                productos_siigo = []
                total_obtenidos = len(df_siigo)
                st.session_state["_loading_progress"] = 100
                st.session_state["_loading_text"] = "Datos cargados desde caché ✓"
                st.session_state.update(
                    {
                        "df_siigo_cache": df_siigo,
                        "productos_siigo_cache": productos_siigo,
                        "total_obtenidos": total_obtenidos,
                        "ultima_actualizacion": ultima_actualizacion,
                    }
                )
            else:
                if "token_siigo" not in st.session_state:
                    st.session_state["_loading_progress"] = 45
                    st.session_state["_loading_text"] = "Autenticando en Siigo..."
                    _render_loading("Autenticando en Siigo…")

                    resultado_siigo = autenticar_siigo(
                        "dirtec@colsabor.com.co",
                        SIIGO_ACCESS_KEY,
                    )
                    if not resultado_siigo["success"]:
                        st.error(
                            "Inicio de sesión correcto en Supabase, pero no fue posible conectar con Siigo."
                        )
                        st.error(resultado_siigo["error"])
                        st.stop()
                    st.session_state["token_siigo"] = resultado_siigo["token"]
                    st.session_state["_loading_progress"] = 55
                    st.session_state["_loading_text"] = "Autenticación exitosa ✓"

                st.session_state["_loading_progress"] = 60
                st.session_state["_loading_text"] = "Obteniendo productos de Siigo..."
                _render_loading("Descargando productos…")

                resultado = obtener_todos_los_productos_siigo(
                    st.session_state["token_siigo"]
                )
                if not resultado["success"]:
                    st.error(resultado["error"])
                    st.stop()
                productos_siigo = resultado["data"]
                total_obtenidos = resultado.get("total", len(productos_siigo))

                st.session_state["_loading_progress"] = 80
                st.session_state["_loading_text"] = "Procesando datos..."
                _render_loading("Procesando datos…")

                df_siigo = procesar_productos_siigo(productos_siigo)
                st.session_state["_loading_progress"] = 95
                st.session_state["_loading_text"] = "Guardando en caché..."

                st.session_state.update(
                    {
                        "df_siigo_cache": df_siigo,
                        "productos_siigo_cache": productos_siigo,
                        "total_obtenidos": total_obtenidos,
                        "ultima_actualizacion": now_colombia(),
                    }
                )
                guardar_productos_siigo(productos_siigo)
                st.session_state["_loading_progress"] = 100
                st.session_state["_loading_text"] = "¡Listo! 🎉"

        # Limpiar flag ANTES del rerun para que tests lo vean eliminado
        if "forzar_actualizacion" in st.session_state:
            del st.session_state["forzar_actualizacion"]
        # Garantizar mínimo 5 segundos de pantalla de carga
        _elapsed = time.time() - _loading_start
        if _elapsed < 5.0:
            time.sleep(5.0 - _elapsed)
        st.rerun()
    else:
        df_excel = st.session_state["df_excel_cache"]
        df_siigo = st.session_state["df_siigo_cache"]
        productos_siigo = st.session_state.get("productos_siigo_cache", [])
        total_obtenidos = st.session_state.get("total_obtenidos", len(df_siigo))

    if "ultima_actualizacion" not in st.session_state:
        st.session_state["ultima_actualizacion"] = now_colombia()

    # ── CRUCE DE INVENTARIOS ──────────────────────────────────────────────────
    df_resultado = cruzar_inventarios(df_excel, df_siigo)

    total = len(df_resultado)
    criticos = len(df_resultado[df_resultado["Estado"].str.contains("Crítico")])
    bajos = len(df_resultado[df_resultado["Estado"].str.contains("Bajo")])
    ok = len(df_resultado[df_resultado["Estado"].str.contains("OK")])
    no_encontrados = len(
        df_resultado[df_resultado["Estado"].str.contains("No encontrado")]
    )

    # ── BENTO MÉTRICO ────────────────────────────────────────────────────────
    st.markdown(
        f"""
        <div class="cs-bento">
          <div class="cs-card cs-card-blue">
            <div class="cs-card-icon">📦</div>
            <div class="cs-card-value">{total}</div>
            <div class="cs-card-label">Total Referencias</div>
            <div class="cs-card-sub">{total_obtenidos} en Siigo</div>
          </div>
          <div class="cs-card cs-card-red">
            <div class="cs-card-icon">🔴</div>
            <div class="cs-card-value">{criticos}</div>
            <div class="cs-card-label">Críticos</div>
            <div class="cs-card-sub">Por debajo del mínimo</div>
          </div>
          <div class="cs-card cs-card-amber">
            <div class="cs-card-icon">🟡</div>
            <div class="cs-card-value">{bajos}</div>
            <div class="cs-card-label">Stock Bajo</div>
            <div class="cs-card-sub">Margen ≤ 20% del mínimo</div>
          </div>
          <div class="cs-card cs-card-green">
            <div class="cs-card-icon">🟢</div>
            <div class="cs-card-value">{ok}</div>
            <div class="cs-card-label">En Orden</div>
            <div class="cs-card-sub">Stock suficiente</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── ANALYTICS ─────────────────────────────────────────────────────────
    st.markdown(
        '<div class="cs-section">📊 Análisis Visual</div>', unsafe_allow_html=True
    )
    ach1, ach2 = st.columns([1, 1.6])
    with ach1:
        st.plotly_chart(
            _build_donut_chart(criticos, bajos, ok, no_encontrados, total),
            use_container_width=True,
            config={"displayModeBar": False},
        )
    with ach2:
        df_deficit = df_resultado[df_resultado["Diferencia"] < 0].nsmallest(
            8, "Diferencia"
        )
        if len(df_deficit) > 0:
            st.plotly_chart(
                _build_deficit_chart(df_deficit),
                use_container_width=True,
                config={"displayModeBar": False},
            )
        else:
            st.markdown(
                '<div class="cs-ok-panel"><span style="font-size:28px">🎯</span>'
                "<div><b>Inventario Saludable</b><br>"
                '<span style="font-size:12px;opacity:.7">Todos los productos en niveles óptimos</span></div></div>',
                unsafe_allow_html=True,
            )

    if no_encontrados > 0:
        st.warning(
            f"⚠️ {no_encontrados} referencia(s) del inventario mínimo no encontradas en Siigo."
        )
        with st.expander(
            f"📋 Ver {no_encontrados} referencias no encontradas", expanded=False
        ):
            df_no_encontradas = df_resultado[
                df_resultado["Estado"].str.contains("No encontrado")
            ][["Referencia", "Nombre", "Mínimo (g)"]].reset_index(drop=True)
            st.table(df_no_encontradas)

    # ── FILTROS ───────────────────────────────────────────────────────────────
    st.markdown('<div class="cs-section">🔍 Filtros</div>', unsafe_allow_html=True)
    st.markdown('<div class="cs-panel">', unsafe_allow_html=True)
    fc1, fc2 = st.columns([1, 1])
    with fc1:
        filtro_estado = st.multiselect(
            "Estado del producto",
            options=["🔴 Crítico", "🟡 Bajo", "🟢 OK", "⚠️ No encontrado en Siigo"],
            default=["🔴 Crítico", "🟡 Bajo"],
        )
    with fc2:
        busqueda = st.text_input(
            "Buscar por referencia o nombre", placeholder="Ej: R003 o AREQUIPE…"
        )
    st.markdown("</div>", unsafe_allow_html=True)

    # Aplicar filtros
    df_filtrado = df_resultado.copy()
    if filtro_estado:
        mascara = df_filtrado["Estado"].apply(
            lambda e: any(
                ("Crítico" in f and "Crítico" in e)
                or ("Bajo" in f and "Bajo" in e and "Crítico" not in e)
                or ("OK" in f and "OK" in e)
                or ("No encontrado" in f and "No encontrado" in e)
                for f in filtro_estado
            )
        )
        df_filtrado = df_filtrado[mascara]
    if busqueda:
        q = busqueda.lower()
        df_filtrado = df_filtrado[
            df_filtrado["Referencia"].str.lower().str.contains(q)
            | df_filtrado["Nombre"].str.lower().str.contains(q)
        ]
    orden_estado = {
        "🔴 Crítico": 0,
        "🟡 Bajo": 1,
        "🟢 OK": 2,
        "⚪ No encontrado en Siigo": 3,
    }
    if not df_filtrado.empty:
        df_filtrado = (
            df_filtrado.assign(_orden=df_filtrado["Estado"].map(orden_estado).fillna(9))
            .sort_values(["_orden", "Diferencia"], ascending=[True, True])
            .drop(columns="_orden")
        )

    # ── TABLA ─────────────────────────────────────────────────────────────────
    st.markdown(
        f'<div class="cs-section">📋 Resultados <span style="font-size:11px;font-weight:400;color:var(--text-muted);text-transform:none;letter-spacing:0">— {len(df_filtrado)} registros</span></div>',
        unsafe_allow_html=True,
    )

    def colorear_estado(val):
        if "Crítico" in str(val):
            return "background-color:#ffcdd2;color:#b71c1c;font-weight:600"
        elif "Bajo" in str(val) and "Crítico" not in str(val):
            return "background-color:#fff9c4;color:#e65100;font-weight:600"
        elif "OK" in str(val):
            return "background-color:#c8e6c9;color:#1b5e20;font-weight:600"
        elif "No encontrado" in str(val):
            return "background-color:#ffccbc;color:#bf360c;font-weight:600"
        return ""

    df_styled = df_filtrado.style.map(colorear_estado, subset=["Estado"]).format(
        {"Mínimo (g)": "{:,.0f}", "Stock Actual": "{:,.0f}", "Diferencia": "{:,.0f}"}
    )
    st.dataframe(df_styled, use_container_width=True, hide_index=True)

    # ── EXPORTAR ──────────────────────────────────────────────────────────────
    df_faltantes = df_resultado[
        df_resultado["Estado"].str.contains("Crítico|Bajo", regex=True)
    ].copy()

    if len(df_faltantes) > 0:
        st.markdown('<div class="cs-section">📥 Exportar</div>', unsafe_allow_html=True)
        _ts = now_colombia().strftime("%Y%m%d_%H%M")
        _ext = "xlsx" if _OPENPYXL_OK else "csv"
        _mime_xl = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        _mime_csv = "text/csv"
        dc1, dc2 = st.columns(2)
        with dc1:
            st.download_button(
                f"📥 Faltantes ({_ext.upper()})",
                data=generar_excel_tabla_descarga(
                    df_faltantes,
                    sheet_title="faltantes",
                    table_display_name="TablaFaltantes",
                ),
                file_name=f"faltantes_{_ts}.{_ext}",
                mime=_mime_xl if _OPENPYXL_OK else _mime_csv,
                use_container_width=True,
                key="download_faltantes_xlsx",
            )
        with dc2:
            st.download_button(
                f"📥 Inventario completo ({_ext.upper()})",
                data=generar_excel_tabla_descarga(
                    df_resultado,
                    sheet_title="inventario",
                    table_display_name="TablaInventario",
                ),
                file_name=f"inventario_{_ts}.{_ext}",
                mime=_mime_xl if _OPENPYXL_OK else _mime_csv,
                use_container_width=True,
                key="download_inventario_completo_xlsx",
            )
        st.markdown(
            f'<div class="cs-export-stats">🔴 {criticos} crítico(s) · 🟡 {bajos} bajo(s) · 📦 {total} referencias totales</div>',
            unsafe_allow_html=True,
        )
    else:
        st.success("🎉 ¡Excelente! No hay productos con stock crítico o bajo.")

    # ── DETALLES TÉCNICOS ─────────────────────────────────────────────────────
    with st.expander("⚙️ Detalles técnicos"):
        st.caption(f"Inventario mínimo: {len(df_excel)} referencias")
        st.caption(f"Productos en Siigo: {len(df_siigo)}")
        st.caption(
            f"Última actualización: {st.session_state['ultima_actualizacion'].strftime('%d/%m/%Y %H:%M:%S')}"
        )
        if len(st.session_state.get("productos_siigo_cache", [])) > 0:
            st.dataframe(df_siigo.head(10), use_container_width=True)

    # ── FOOTER ────────────────────────────────────────────────────────────────
    st.markdown(
        """
        <div style="
          position:fixed;bottom:0;left:0;right:0;z-index:900;
          padding:10px 24px;
          background:var(--nav-bg);
          border-top:1px solid var(--nav-border);
          backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);
          display:flex;align-items:center;justify-content:center;
        ">
          <p style="color:var(--text-muted);font-size:11px;letter-spacing:0.06em;
            margin:0;font-family:'Inter',sans-serif;font-weight:500">
            COLSABOR S.A.S &nbsp;·&nbsp; Monitor de Inventario &nbsp;·&nbsp; © 2026
          </p>
        </div>
        """,
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
