from __future__ import annotations

import io
import os
import tomllib
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from supabase import create_client


SIIGO_API_BASE_URL = "https://api.siigo.com/v1"
SIIGO_ACCESS_KEY = "MmQzMDk0NjYtZjc3Ny00YzU0LWFmNDMtMjhiYzcxNGM5NTBhOnoyeTk5KE4uYkc="
SIIGO_USERNAME = "dirtec@colsabor.com.co"
TABLE_SIIGO_CACHE = "siigo_products_cache"
COL_TZ = ZoneInfo("America/Bogota")

_SUPABASE_URL_DEFAULT = "https://uinqrkxlkjowixmtzold.supabase.co"
_SUPABASE_KEY_DEFAULT = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InVpbnFya3hsa2pvd2l4bXR6b2xkIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzYyMTg2OTgsImV4cCI6MjA5MTc5NDY5OH0"
    ".ifGRvqwAtI-6D72_BC7uih-88boy2wcseBUEi-o_0ek"
)

ALLOWED_EMAILS = {
    "dirtec@colsabor.com.co",
    "gerencia@colsabor.com.co",
    "samuelrestrepodev@gmail.com",
}


def now_colombia() -> datetime:
    return datetime.now(COL_TZ)


def _load_supabase_from_local_secrets() -> tuple[str, str]:
    candidate_paths = [
        Path.cwd() / ".streamlit" / "secrets.toml",
        Path(__file__).resolve().parent / ".streamlit" / "secrets.toml",
        Path(__file__).resolve().parents[1] / ".streamlit" / "secrets.toml",
    ]
    for secrets_path in candidate_paths:
        if not secrets_path.exists():
            continue
        data = tomllib.loads(secrets_path.read_text(encoding="utf-8"))
        url = str(data.get("SUPABASE_URL", "") or "").strip()
        key = str(data.get("SUPABASE_KEY", "") or "").strip()
        if url and key:
            return url, key
    return "", ""


def get_supabase_settings() -> tuple[str, str]:
    local_url, local_key = _load_supabase_from_local_secrets()
    return (
        os.getenv("SUPABASE_URL") or local_url or _SUPABASE_URL_DEFAULT,
        os.getenv("SUPABASE_KEY") or local_key or _SUPABASE_KEY_DEFAULT,
    )


def get_supabase_client():
    url, key = get_supabase_settings()
    return create_client(url, key)


def authenticate_user(email: str, password: str) -> dict[str, Any]:
    email_clean = email.strip().lower()
    if email_clean not in ALLOWED_EMAILS:
        return {"success": False, "error": "Este correo no esta autorizado."}
    try:
        response = get_supabase_client().auth.sign_in_with_password(
            {"email": email_clean, "password": password}
        )
    except Exception as exc:
        return {"success": False, "error": str(exc)}
    if not response.session:
        return {"success": False, "error": "Credenciales incorrectas."}
    return {
        "success": True,
        "email": response.user.email if response.user else email_clean,
        "access_token": response.session.access_token,
    }


def autenticar_siigo(username: str = SIIGO_USERNAME, access_key: str = SIIGO_ACCESS_KEY) -> dict:
    headers = {"Content-Type": "application/json", "Partner-Id": "ColsaborApp"}
    payload = {"username": username, "access_key": access_key}
    try:
        response = requests.post("https://api.siigo.com/auth", json=payload, headers=headers, timeout=30)
        if response.status_code == 200:
            return {"success": True, "token": response.json().get("access_token")}
        return {"success": False, "error": f"Error de autenticacion: {response.status_code} - {response.text}"}
    except requests.exceptions.RequestException as exc:
        return {"success": False, "error": f"Error de conexion: {exc}"}


def obtener_todos_los_productos_siigo(token: str) -> dict:
    headers = {"Authorization": token, "Content-Type": "application/json", "Partner-Id": "ColsaborApp"}
    todos_productos = []
    page = 1
    page_size = 100
    try:
        while True:
            response = requests.get(
                f"{SIIGO_API_BASE_URL}/products",
                headers=headers,
                params={"page": page, "page_size": page_size},
                timeout=60,
            )
            if response.status_code != 200:
                return {"success": False, "error": f"Error al obtener productos: {response.status_code} - {response.text}"}
            data = response.json()
            productos_pagina = data if isinstance(data, list) else data.get("results", [])
            if not productos_pagina:
                break
            todos_productos.extend(productos_pagina)
            if len(productos_pagina) < page_size:
                break
            page += 1
    except requests.exceptions.RequestException as exc:
        return {"success": False, "error": f"Error de conexion: {exc}"}
    return {"success": True, "data": todos_productos, "total": len(todos_productos)}


def guardar_productos_siigo(productos_siigo: list) -> bool:
    try:
        get_supabase_client().table(TABLE_SIIGO_CACHE).upsert(
            {"id": 1, "data": productos_siigo, "updated_at": now_colombia().isoformat()},
            on_conflict="id",
        ).execute()
        return True
    except Exception:
        return False


def cargar_productos_siigo_guardados(max_age_hours: int = 24):
    try:
        response = get_supabase_client().table(TABLE_SIIGO_CACHE).select("*").eq("id", 1).execute()
        if not response.data:
            return None
        row = response.data[0]
        updated_at_str = str(row["updated_at"]).replace("Z", "+00:00")
        updated_at = datetime.fromisoformat(updated_at_str)
        if updated_at.tzinfo is None:
            updated_at = updated_at.replace(tzinfo=COL_TZ)
        updated_at = updated_at.astimezone(COL_TZ)
        if (now_colombia() - updated_at).total_seconds() / 3600 > max_age_hours:
            return None
        return procesar_productos_siigo(row["data"]), row["data"], updated_at
    except Exception:
        return None


def cargar_inventario_minimo_supabase() -> pd.DataFrame | None:
    try:
        response = get_supabase_client().table("inventario_minimo").select("*").execute()
        if not response.data:
            return None
        df = pd.DataFrame(response.data)
        df = df.rename(columns={"codigo": "referencia", "nombre": "nombre", "inv_minimo": "inventario_minimo"})
        return normalize_inventory_dataframe(df)
    except Exception:
        return None


def parse_inventory_excel(src: str | bytes | io.BytesIO) -> pd.DataFrame:
    df = pd.read_excel(src)
    return normalize_inventory_dataframe(df)


def normalize_inventory_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    normalized = df.copy()
    lookup = {str(col).strip().lower(): col for col in normalized.columns}
    ref_col = lookup.get("referencia") or lookup.get("codigo") or lookup.get("código")
    name_col = lookup.get("nombre")
    min_col = (
        lookup.get("inventario mínimo por gramos")
        or lookup.get("inventario minimo por gramos")
        or lookup.get("inventario_minimo")
        or lookup.get("inv_minimo")
        or lookup.get("mínimo (g)")
        or lookup.get("minimo (g)")
    )
    missing = [
        label
        for label, value in {"referencia/codigo": ref_col, "nombre": name_col, "inventario minimo": min_col}.items()
        if value is None
    ]
    if missing:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(missing)}")

    out = normalized[[ref_col, name_col, min_col]].copy()
    out.columns = ["referencia", "nombre", "inventario_minimo"]
    out["referencia"] = out["referencia"].astype(str).str.strip()
    out["nombre"] = out["nombre"].astype(str).str.strip()
    out["inventario_minimo"] = pd.to_numeric(out["inventario_minimo"], errors="coerce").fillna(0)
    return out[out["referencia"] != ""].reset_index(drop=True)


def procesar_productos_siigo(productos: list) -> pd.DataFrame:
    datos = []
    for producto in productos:
        referencia = producto.get("code", "")
        nombre = producto.get("name", "")
        if not referencia or not nombre:
            continue
        stock_actual = 0.0
        if "available_quantity" in producto:
            stock_actual = float(producto["available_quantity"] or 0)
        elif "stock" in producto:
            stock_actual = float(producto["stock"] or 0)
        elif "warehouses" in producto and isinstance(producto["warehouses"], list):
            stock_actual = sum(float(bodega.get("quantity", 0) or 0) for bodega in producto["warehouses"])
        datos.append(
            {
                "referencia_siigo": str(referencia).strip(),
                "nombre_siigo": str(nombre).strip(),
                "stock_actual": stock_actual,
            }
        )
    if not datos:
        return pd.DataFrame(columns=["referencia_siigo", "nombre_siigo", "stock_actual"])
    return pd.DataFrame(datos)


def cruzar_inventarios(df_excel: pd.DataFrame, df_siigo: pd.DataFrame) -> pd.DataFrame:
    df_cruzado = df_excel.merge(df_siigo, left_on="referencia", right_on="referencia_siigo", how="left")
    df_siigo_aux = df_siigo.copy()
    df_siigo_aux["nombre_normalizado"] = (
        df_siigo_aux["nombre_siigo"].astype(str).str.upper().str.strip().str.replace(r"\s+", " ", regex=True)
    )
    no_encontrados_mask = df_cruzado["referencia_siigo"].isna()
    if no_encontrados_mask.any():
        df_cruzado["nombre_normalizado"] = (
            df_cruzado["nombre"].astype(str).str.upper().str.strip().str.replace(r"\s+", " ", regex=True)
        )
        for idx in df_cruzado[no_encontrados_mask].index:
            nombre_norm = df_cruzado.loc[idx, "nombre_normalizado"]
            match = df_siigo_aux[df_siigo_aux["nombre_normalizado"] == nombre_norm]
            if len(match) > 0:
                match_idx = match.index[0]
                df_cruzado.loc[idx, "referencia_siigo"] = df_siigo_aux.loc[match_idx, "referencia_siigo"]
                df_cruzado.loc[idx, "nombre_siigo"] = df_siigo_aux.loc[match_idx, "nombre_siigo"]
                df_cruzado.loc[idx, "stock_actual"] = df_siigo_aux.loc[match_idx, "stock_actual"]

    df_cruzado["encontrado_en_siigo"] = df_cruzado["referencia_siigo"].notna()
    df_cruzado["stock_actual"] = df_cruzado["stock_actual"].fillna(0)
    df_cruzado["diferencia"] = df_cruzado["stock_actual"] - df_cruzado["inventario_minimo"]

    def determinar_estado(row):
        if not row["encontrado_en_siigo"]:
            return "No encontrado en Siigo"
        if row["stock_actual"] < row["inventario_minimo"]:
            return "Crítico"
        if row["stock_actual"] <= row["inventario_minimo"] * 1.2:
            return "Bajo"
        return "OK"

    df_cruzado["estado"] = df_cruzado.apply(determinar_estado, axis=1)
    df_resultado = df_cruzado[["referencia", "nombre", "inventario_minimo", "stock_actual", "diferencia", "estado"]].copy()
    df_resultado.columns = ["Referencia", "Nombre", "Mínimo (g)", "Stock Actual", "Diferencia", "Estado"]
    return df_resultado


def generar_excel_tabla_descarga(
    df: pd.DataFrame,
    *,
    sheet_title: str = "Datos",
    table_display_name: str = "TablaInventario",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31].replace("/", "-")
    ncols = max(1, len(df.columns))
    header_fill = PatternFill("solid", fgColor="1F6F4F")
    header_font = Font(color="FFFFFF", bold=True)
    align = Alignment(vertical="center", wrap_text=True)
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    last_col = get_column_letter(ncols)
    tab = Table(displayName=table_display_name[:255], ref=f"A1:{last_col}{len(df) + 1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    for i, col in enumerate(df.columns, 1):
        letter = get_column_letter(i)
        sample = [str(col)] + [str(x) for x in df.iloc[:, i - 1].head(50).tolist()]
        ws.column_dimensions[letter].width = min(48, max(10, max(len(s) for s in sample) + 2))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_siigo_dataframe(refresh: bool = False) -> tuple[pd.DataFrame, int, datetime, str]:
    cached = None if refresh else cargar_productos_siigo_guardados()
    if cached is not None:
        df_siigo, productos, updated_at = cached
        return df_siigo, len(df_siigo), updated_at, "cache"
    auth = autenticar_siigo()
    if not auth["success"]:
        raise RuntimeError(auth["error"])
    result = obtener_todos_los_productos_siigo(auth["token"])
    if not result["success"]:
        raise RuntimeError(result["error"])
    productos = result["data"]
    guardar_productos_siigo(productos)
    return procesar_productos_siigo(productos), int(result.get("total", len(productos))), now_colombia(), "siigo"


def build_inventory_payload(df_excel: pd.DataFrame, df_siigo: pd.DataFrame, total_siigo: int, updated_at: datetime, source: str) -> dict[str, Any]:
    df_resultado = cruzar_inventarios(df_excel, df_siigo)
    total = len(df_resultado)
    criticos = int(df_resultado["Estado"].str.contains("Crítico").sum())
    bajos = int(df_resultado["Estado"].str.contains("Bajo").sum())
    ok = int((df_resultado["Estado"] == "OK").sum())
    no_encontrados = int(df_resultado["Estado"].str.contains("No encontrado").sum())
    deficits = df_resultado[df_resultado["Diferencia"] < 0].nsmallest(12, "Diferencia")
    missing = df_resultado[df_resultado["Estado"].str.contains("No encontrado")]
    return {
        "summary": {
            "total": total,
            "criticos": criticos,
            "bajos": bajos,
            "ok": ok,
            "no_encontrados": no_encontrados,
            "total_siigo": total_siigo,
            "ultima_actualizacion": updated_at.isoformat(),
        },
        "rows": df_resultado.to_dict(orient="records"),
        "missing": missing.to_dict(orient="records"),
        "deficits": deficits.to_dict(orient="records"),
        "source": {"inventory": "supabase", "siigo": source},
    }


def get_current_inventory_payload(refresh: bool = False) -> dict[str, Any]:
    df_excel = cargar_inventario_minimo_supabase()
    if df_excel is None:
        raise RuntimeError("No se pudo cargar inventario minimo desde Supabase.")
    df_siigo, total_siigo, updated_at, source = get_siigo_dataframe(refresh=refresh)
    return build_inventory_payload(df_excel, df_siigo, total_siigo, updated_at, source)
